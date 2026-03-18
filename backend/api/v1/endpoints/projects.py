from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Response, Form, Body
from datetime import date, timedelta, datetime
from typing import Optional
import os
import asyncio
import logging
from uuid import uuid4
import csv
import io
from pydantic import BaseModel

logger = logging.getLogger(__name__)

from backend.core.deps import DBSessionDep, require_roles, get_current_user
from backend.iam.decorators import require_permission
from backend.core.config import settings
from backend.core.security import verify_password
from backend.repositories.project_repository import ProjectRepository
from backend.repositories.transaction_repository import TransactionRepository
from backend.repositories.user_repository import UserRepository
from backend.schemas.project import ProjectCreate, ProjectOut, ProjectUpdate
from backend.schemas.recurring_transaction import RecurringTransactionTemplateCreate
from backend.services.project_service import ProjectService, calculate_monthly_income_amount
from backend.services.recurring_transaction_service import RecurringTransactionService
from backend.services.financial_aggregation_service import FinancialAggregationService
from backend.services.budget_service import BudgetService
from backend.services.fund_service import FundService
from backend.services.s3_service import S3Service
from backend.services.audit_service import AuditService
from backend.services.contract_period_service import ContractPeriodService
from backend.models.user import UserRole
from backend.models.project import Project
from backend.models.subproject import Subproject
from backend.models.recurring_transaction import RecurringTransactionTemplate
from backend.models.fund import Fund
from backend.models.budget import Budget
from backend.models.contract_period import ContractPeriod
from backend.models.quote_project import QuoteProject
from sqlalchemy import delete, update

router = APIRouter()


def get_uploads_dir() -> str:
    """Get absolute path to uploads directory, resolving relative paths relative to backend directory"""
    if os.path.isabs(settings.FILE_UPLOAD_DIR):
        return settings.FILE_UPLOAD_DIR
    else:
        # Get the directory where this file is located, then go up to backend directory
        current_dir = os.path.dirname(os.path.abspath(__file__))
        # Go from api/v1/endpoints to backend directory
        backend_dir = os.path.dirname(os.path.dirname(os.path.dirname(current_dir)))
        return os.path.abspath(os.path.join(backend_dir, settings.FILE_UPLOAD_DIR))


async def _build_projects_list(db, projects) -> list[ProjectOut]:
    """Shared helper: batch-load funds & earliest contract dates, then build ProjectOut list.
    Replaces the former N+1 loop (2 queries per project) with 2 batch queries total."""
    from sqlalchemy import select, func
    from backend.models.fund import Fund
    from backend.models.contract_period import ContractPeriod

    if not projects:
        return []

    project_ids = [p.id for p in projects]

    # Batch query 1: all funds for these projects (single query)
    funds_result = await db.execute(
        select(Fund).where(Fund.project_id.in_(project_ids))
    )
    funds_by_project = {f.project_id: f for f in funds_result.scalars().all()}

    # Batch query 2: earliest start_date per project (single query)
    earliest_result = await db.execute(
        select(ContractPeriod.project_id, func.min(ContractPeriod.start_date))
        .where(ContractPeriod.project_id.in_(project_ids))
        .group_by(ContractPeriod.project_id)
    )
    earliest_by_project = {row[0]: row[1] for row in earliest_result.all()}

    result = []
    for project in projects:
        fund = funds_by_project.get(project.id)
        first_start = earliest_by_project.get(project.id)
        if first_start is None and project.start_date:
            s = project.start_date
            first_start = s.date() if hasattr(s, 'date') else s

        # Create ProjectOut instance using model_validate to ensure proper validation
        # Keep dates as date objects (not strings) - Pydantic will serialize them
        project_data = {
            "id": project.id,
            "name": project.name,
            "description": project.description,
            "start_date": project.start_date,
            "end_date": project.end_date,
            "contract_duration_months": project.contract_duration_months,
            "budget_monthly": float(project.budget_monthly) if project.budget_monthly else 0.0,
            "budget_annual": float(project.budget_annual) if project.budget_annual else 0.0,
            "manager_id": project.manager_id,
            "relation_project": project.relation_project,
            "is_parent_project": project.is_parent_project,
            "show_in_quotes_tab": getattr(project, 'show_in_quotes_tab', False),
            "is_active": project.is_active,
            "num_residents": project.num_residents,
            "monthly_price_per_apartment": float(project.monthly_price_per_apartment) if project.monthly_price_per_apartment else None,
            "address": project.address,
            "city": project.city,
            "image_url": project.image_url,
            "contract_file_url": project.contract_file_url,
            "is_active": project.is_active,
            "created_at": project.created_at,
            "total_value": float(getattr(project, 'total_value', 0.0)),
            "has_fund": fund is not None,
            "monthly_fund_amount": float(fund.monthly_amount) if fund else None,
            "first_contract_start_date": first_start.isoformat() if first_start else None,
        }
        project_out = ProjectOut.model_validate(project_data)
        result.append(project_out)

    return result


@router.get("/", response_model=list[ProjectOut])
async def list_projects(db: DBSessionDep, include_archived: bool = Query(False), only_archived: bool = Query(False), user = Depends(get_current_user)):
    """List projects - accessible to all authenticated users"""
    projects = await ProjectRepository(db).list(include_archived=include_archived, only_archived=only_archived)
    return await _build_projects_list(db, projects)

@router.get("", response_model=list[ProjectOut])
async def list_projects_no_slash(db: DBSessionDep, include_archived: bool = Query(False), only_archived: bool = Query(False), user = Depends(get_current_user)):
    """Alias without trailing slash to avoid 404 when redirect_slashes=False"""
    projects = await ProjectRepository(db).list(include_archived=include_archived, only_archived=only_archived)
    return await _build_projects_list(db, projects)

@router.get("/check-name")
async def check_project_name(
    db: DBSessionDep,
    name: str = Query(..., description="Project name to check"),
    exclude_id: Optional[int] = Query(None, description="Project ID to exclude from check (for updates)"),
    user=Depends(get_current_user),
):
    """Check if a project name already exists (compare trimmed). For create/edit any project."""
    name_trimmed = (name or "").strip()
    if not name_trimmed:
        return {"exists": False, "available": False}
    existing = await ProjectRepository(db).get_project_by_name(
        name_trimmed, exclude_project_id=exclude_id
    )
    return {
        "exists": existing is not None,
        "available": existing is None,
    }


@router.get("/check-parent-name")
async def check_parent_project_name(
    db: DBSessionDep,
    name: str = Query(..., description="Parent project name to check"),
    exclude_id: Optional[int] = Query(None, description="Project ID to exclude (for edit)"),
    user=Depends(get_current_user),
):
    """Check if a parent project with this name already exists. For use in create/edit parent project modal."""
    name_trimmed = (name or "").strip()
    if not name_trimmed:
        return {"available": False, "reason": "empty"}
    existing = await ProjectRepository(db).get_parent_project_by_name(
        name_trimmed, exclude_project_id=exclude_id
    )
    return {"available": existing is None}


@router.get("/profitability-alerts")
async def get_profitability_alerts(
    db: DBSessionDep,
    user = Depends(get_current_user)
):
    """
    Get projects and sub-projects with profitability issues based on last 6 months of data.
    Returns projects with profit margin <= -10% (loss-making projects).
    
    OPTIMIZED: Fetches all data in 2 queries instead of N*3 queries per project.
    """
    from sqlalchemy import select, and_
    from backend.models.project import Project
    from backend.models.transaction import Transaction
    from datetime import timedelta

    # Calculate dates
    today = date.today()
    six_months_ago = today - timedelta(days=180)
    one_year_ago = today - timedelta(days=365)

    # OPTIMIZED: Get all projects in ONE query
    projects_result = await db.execute(select(Project))
    all_projects = list(projects_result.scalars().all())
    
    if not all_projects:
        return {"alerts": []}
    
    # Extract project IDs
    project_ids = [p.id for p in all_projects]
    
    # OPTIMIZED: Get ALL transactions for ALL projects in ONE query (last year)
    all_transactions_query = select(Transaction).where(
        and_(
            Transaction.project_id.in_(project_ids),
            Transaction.tx_date >= one_year_ago,
            Transaction.tx_date <= today
        )
    )
    all_transactions_result = await db.execute(all_transactions_query)
    all_transactions = list(all_transactions_result.scalars().all())
    
    # Group transactions by project_id for efficient lookup
    transactions_by_project = {}
    for tx in all_transactions:
        if tx.project_id not in transactions_by_project:
            transactions_by_project[tx.project_id] = []
        transactions_by_project[tx.project_id].append(tx)

    alerts = []

    # Process each project FROM MEMORY (no more DB queries!)
    for project in all_projects:
        project_transactions = transactions_by_project.get(project.id, [])
        
        if not project_transactions:
            continue
        
        # Filter transactions for last 6 months
        transactions_6m = [
            t for t in project_transactions 
            if t.tx_date >= six_months_ago
        ]
        
        # Calculate income and expenses (exclude fund transactions)
        if transactions_6m:
            income = sum(
                float(t.amount) for t in transactions_6m 
                if t.type == 'Income' and not getattr(t, 'from_fund', False)
            )
            expense = sum(
                float(t.amount) for t in transactions_6m 
                if t.type == 'Expense' and not getattr(t, 'from_fund', False)
            )
        else:
            # No transactions in 6 months - use all transactions from last year
            income = sum(
                float(t.amount) for t in project_transactions 
                if t.type == 'Income' and not getattr(t, 'from_fund', False)
            )
            expense = sum(
                float(t.amount) for t in project_transactions 
                if t.type == 'Expense' and not getattr(t, 'from_fund', False)
            )
        
        profit = income - expense

        # Calculate profit margin
        if income > 0:
            profit_margin = (profit / income) * 100
        elif expense > 0:
            profit_margin = -100
        else:
            continue

        # Only include projects with profit margin <= -10% (loss-making)
        if profit_margin <= -10:
            is_subproject = project.relation_project is not None

            alerts.append({
                'id': int(project.id),
                'name': str(project.name),
                'profit_margin': float(round(profit_margin, 1)),
                'income': float(income),
                'expense': float(expense),
                'profit': float(profit),
                'is_subproject': bool(is_subproject),
                'parent_project_id': int(project.relation_project) if project.relation_project else None
            })

    # Sort by profit margin (most negative first)
    alerts.sort(key=lambda x: x['profit_margin'])

    result = {
        'alerts': alerts,
        'count': int(len(alerts)),
        'period_start': str(six_months_ago.isoformat()),
        'period_end': str(today.isoformat())
    }

    return result

@router.get("/{project_id}", response_model=ProjectOut)
async def get_project(project_id: int, db: DBSessionDep, user = Depends(get_current_user)):
    """Get project details - accessible to all authenticated users"""
    project = await ProjectRepository(db).get_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Add fund information to project response
    from backend.services.fund_service import FundService
    from backend.repositories.contract_period_repository import ContractPeriodRepository
    fund_service = FundService(db)
    fund = await fund_service.get_fund_by_project(project_id)
    period_repo = ContractPeriodRepository(db)
    first_start = await period_repo.get_earliest_start_date(project_id)
    if first_start is None and project.start_date:
        s = project.start_date
        first_start = s.date() if hasattr(s, 'date') else s

    
    # Create ProjectOut instance using model_validate to ensure proper validation
    # Keep dates as date objects (not strings) - Pydantic will serialize them
    project_data = {
        "id": project.id,
        "name": project.name,
        "description": project.description,
        "start_date": project.start_date,
        "end_date": project.end_date,
        "contract_duration_months": project.contract_duration_months,
        "budget_monthly": float(project.budget_monthly) if project.budget_monthly else 0.0,
        "budget_annual": float(project.budget_annual) if project.budget_annual else 0.0,
        "manager_id": project.manager_id,
        "relation_project": project.relation_project,
        "is_parent_project": project.is_parent_project,
        "show_in_quotes_tab": getattr(project, 'show_in_quotes_tab', False),
        "is_active": project.is_active,
        "num_residents": project.num_residents,
        "monthly_price_per_apartment": float(project.monthly_price_per_apartment) if project.monthly_price_per_apartment else None,
        "address": project.address,
        "city": project.city,
        "image_url": project.image_url,
        "contract_file_url": project.contract_file_url,
        "is_active": project.is_active,
        "created_at": project.created_at,
        "total_value": float(getattr(project, 'total_value', 0.0)),
        "has_fund": fund is not None,
        "monthly_fund_amount": float(fund.monthly_amount) if fund else None,
        "first_contract_start_date": first_start.isoformat() if first_start else None,
    }
    return ProjectOut.model_validate(project_data)

@router.get("/{project_id}/full")
async def get_project_full(
    project_id: int, 
    db: DBSessionDep, 
    user = Depends(get_current_user),
    period_id: Optional[int] = None
):
    """
    OPTIMIZED: Get complete project data in a single API call.
    Returns: project info + transactions + budgets + expense categories + fund data
    
    This replaces 5+ separate API calls with ONE, dramatically improving page load time.
    
    Optional period_id parameter: When provided, returns data filtered to that specific
    contract period (for viewing historical periods).
    """
    from sqlalchemy import select, and_, or_
    from sqlalchemy.orm import selectinload
    from backend.models.transaction import Transaction
    from backend.models.budget import Budget
    from backend.models.fund import Fund
    from backend.models.unforeseen_transaction import UnforeseenTransaction
    from backend.services.budget_service import BudgetService
    from backend.services.report_service import ReportService
    from backend.repositories.contract_period_repository import ContractPeriodRepository
    from backend.repositories.quote_project_repository import QuoteProjectRepository
    
    # Get project
    project = await ProjectRepository(db).get_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # If period_id is provided, load the specific period and filter data accordingly
    selected_period = None
    period_start_date = None
    period_end_date = None
    
    if period_id:
        period_repo = ContractPeriodRepository(db)
        selected_period = await period_repo.get_by_id(period_id)
        if not selected_period:
            raise HTTPException(status_code=404, detail="Contract period not found")
        if selected_period.project_id != project_id:
            raise HTTPException(status_code=400, detail="Contract period does not belong to this project")
        period_start_date = selected_period.start_date
        period_end_date = selected_period.end_date
    
    # Batch query all data in parallel-style (minimize round trips)
    # Query 1: Get transactions for project with category info
    # If period_id provided, filter by period dates
    if period_id and period_start_date and period_end_date:
        # Filter transactions: same logic as current contract -
        # regular within [start,end), period overlap, fund tx within [start,end)
        transactions_query = select(Transaction).options(
            selectinload(Transaction.category)
        ).where(
            and_(
                Transaction.project_id == project_id,
                or_(
                    # Fund transactions within date range (like current contract)
                    and_(
                        Transaction.from_fund == True,
                        Transaction.tx_date >= period_start_date,
                        Transaction.tx_date < period_end_date
                    ),
                    # Regular (non-fund) transactions within date range [start_date, end_date)
                    and_(
                        Transaction.from_fund == False,
                        Transaction.tx_date >= period_start_date,
                        Transaction.tx_date < period_end_date,
                        or_(
                            Transaction.period_start_date.is_(None),
                            Transaction.period_end_date.is_(None)
                        )
                    ),
                    # Period transactions that overlap with the period
                    and_(
                        Transaction.period_start_date.is_not(None),
                        Transaction.period_end_date.is_not(None),
                        Transaction.period_start_date < period_end_date,
                        Transaction.period_end_date >= period_start_date
                    )
                )
            )
        ).order_by(Transaction.tx_date.desc())
    else:
        transactions_query = select(Transaction).options(
            selectinload(Transaction.category)
        ).where(
            Transaction.project_id == project_id
        ).order_by(Transaction.tx_date.desc())
    
    # Query 2: Get all active budgets for project
    budgets_query = select(Budget).where(
        and_(
            Budget.project_id == project_id,
            Budget.is_active == True
        )
    )
    
    # Query 3: Get fund for project
    funds_query = select(Fund).where(Fund.project_id == project_id)
    
    # Query 4: Get contract periods and current period
    from backend.services.contract_period_service import ContractPeriodService
    contract_service = ContractPeriodService(db)
    
    # Ensure project is up to date with current date (catch up if needed)
    # Only run if the project might need renewal (end_date has passed or is today)
    today = date.today()
    if project.end_date and project.end_date <= today:
        await contract_service.check_and_renew_contract(project_id)
    
    # Also ensure all recurring transactions are generated for the current state of the project
    # This is important when new contract periods were just created by check_and_renew_contract
    # Only run if the project has recurring templates and might have missing transactions
    from backend.services.recurring_transaction_service import RecurringTransactionService
    recurring_service = RecurringTransactionService(db)
    await recurring_service.ensure_project_transactions_generated(project_id)
    
    # Refresh project to get updated dates
    await db.refresh(project)
    
    # Execute all queries
    tx_result, budgets_result, fund_result, current_period, contract_periods_by_year = await asyncio.gather(
        db.execute(transactions_query),
        db.execute(budgets_query),
        db.execute(funds_query),
        contract_service.get_current_contract_period(project_id),
        contract_service.get_previous_contracts_by_year(project_id),
        return_exceptions=True
    )
    
    # Handle exceptions from gather
    if isinstance(current_period, Exception):
        logger.exception("Failed to get current_period for project %s", project_id, exc_info=current_period)
        current_period = None
    if isinstance(contract_periods_by_year, Exception):
        logger.exception("Failed to get contract_periods_by_year for project %s", project_id, exc_info=contract_periods_by_year)
        contract_periods_by_year = {}
    
    # Convert contract_periods_by_year to list format
    periods_by_year_list = []
    for year in sorted(contract_periods_by_year.keys(), reverse=True):
        periods_by_year_list.append({
            'year': year,
            'periods': contract_periods_by_year[year]
        })
    
    # Process transactions
    transactions_list = []
    expense_by_category = {}
    res = await db.execute(
        select(UnforeseenTransaction.resulting_transaction_id).where(
            and_(
                UnforeseenTransaction.project_id == project_id,
                UnforeseenTransaction.resulting_transaction_id.isnot(None)
            )
        )
    )
    unforeseen_resulting_ids = {r[0] for r in res.all() if r[0] is not None}
    if not isinstance(tx_result, Exception):
        transactions = list(tx_result.scalars().all())
        for tx in transactions:
            category_name = tx.category.name if tx.category else None
            tx_dict = {
                "id": tx.id,
                "project_id": tx.project_id,
                "tx_date": tx.tx_date.isoformat() if tx.tx_date else None,
                "type": tx.type,
                "amount": float(tx.amount) if tx.amount else 0,
                "description": tx.description,
                "category": category_name,
                "payment_method": tx.payment_method,
                "notes": tx.notes,
                "is_exceptional": getattr(tx, 'is_exceptional', False),
                "is_generated": getattr(tx, 'is_generated', False),
                "supplier_id": tx.supplier_id,
                "from_fund": getattr(tx, 'from_fund', False),
                "file_path": getattr(tx, 'file_path', None),
                "recurring_template_id": getattr(tx, 'recurring_template_id', None),
                "period_start_date": tx.period_start_date.isoformat() if getattr(tx, 'period_start_date', None) else None,
                "period_end_date": tx.period_end_date.isoformat() if getattr(tx, 'period_end_date', None) else None,
                "created_by_user_id": getattr(tx, 'created_by_user_id', None),
                "is_unforeseen": tx.id in unforeseen_resulting_ids
            }
            transactions_list.append(tx_dict)
            
            # Calculate expense categories
            if tx.type == 'Expense' and category_name and not getattr(tx, 'from_fund', False):
                expense_by_category[category_name] = expense_by_category.get(category_name, 0) + float(tx.amount or 0)
    
    # Process budgets with spending calculations
    # Filter budgets by period: when viewing a specific period, show only that period's budgets
    budgets_list = []
    if not isinstance(budgets_result, Exception):
        budgets = list(budgets_result.scalars().all())
        budget_service = BudgetService(db)
        
        # Determine which period's budgets to show
        effective_period_id = period_id
        if effective_period_id is None:
            # When viewing current period, try to get the current period's ID
            if current_period and isinstance(current_period, dict):
                effective_period_id = current_period.get("period_id")
            elif current_period and hasattr(current_period, 'get'):
                # Handle case where current_period might be a different dict-like object
                effective_period_id = current_period.get("period_id") if callable(getattr(current_period, 'get', None)) else None
            elif current_period and hasattr(current_period, 'id'):
                # Handle case where current_period is an object with id attribute
                effective_period_id = current_period.id
            
            # Fallback: if we still don't have a period_id, try to get it directly from the database
            if effective_period_id is None and project and project.start_date:
                try:
                    period_repo = ContractPeriodRepository(db)
                    all_periods = await period_repo.get_by_project(project_id)
                    for period in all_periods:
                        if period.start_date == project.start_date:
                            effective_period_id = period.id
                            logger.debug("Found current period ID %s from DB for project %s", effective_period_id, project_id)
                            break
                except Exception as e:
                    logger.warning("Error getting period from DB for project %s: %s", project_id, e)
        
        # Debug logging
        total_budgets_before_filter = len(budgets)
        budgets_by_period = {}
        for b in budgets:
            period_id_val = getattr(b, "contract_period_id", None)
            budgets_by_period[period_id_val] = budgets_by_period.get(period_id_val, 0) + 1
        logger.debug("Project %s: %d total budgets, effective_period_id=%s, budgets by period: %s", project_id, total_budgets_before_filter, effective_period_id, budgets_by_period)
        
        # Filter to this period's budgets (contract_period_id match)
        # If effective_period_id is set, show budgets for that specific period
        # If no budgets found for that period, fallback to budgets with NULL contract_period_id (old budgets)
        # If effective_period_id is None, show budgets with NULL contract_period_id (old budgets without period assignment)
        if effective_period_id is not None:
            # First try to get budgets for the specific period
            period_budgets = [b for b in budgets if getattr(b, "contract_period_id", None) == effective_period_id]
            if period_budgets:
                budgets = period_budgets
                logger.debug("Filtered to %d budgets for period_id=%s", len(budgets), effective_period_id)
            else:
                # No budgets for this period, fallback to NULL budgets (old budgets without period assignment)
                null_budgets = [b for b in budgets if getattr(b, "contract_period_id", None) is None]
                if null_budgets:
                    budgets = null_budgets
                    logger.warning("No budgets for period_id=%s, using %d budgets with NULL contract_period_id as fallback", effective_period_id, len(budgets))
                else:
                    # If no null budgets either, keep all budgets as last resort
                    logger.warning("No budgets for period_id=%s and no NULL budgets, showing all %d budgets as fallback", effective_period_id, len(budgets))
        else:
            # When viewing current period but no period_id found:
            # Show budgets with NULL contract_period_id (old budgets)
            # If none found, show all budgets as fallback
            null_budgets = [b for b in budgets if getattr(b, "contract_period_id", None) is None]
            if null_budgets:
                budgets = null_budgets
                logger.debug("Using %d budgets with NULL contract_period_id (no period_id found)", len(budgets))
            else:
                # If no null budgets, keep all budgets (they might all be assigned to periods)
                # This ensures budgets are visible even if period detection fails
                logger.warning("No NULL budgets found, showing all %d budgets as fallback", len(budgets))
        for budget in budgets:
            # Calculate spent amount from already-loaded transactions for this budget's category
            spent = sum(
                tx["amount"] for tx in transactions_list
                if tx["type"] == "Expense"
                and tx.get("category") == budget.category
                and not tx.get("from_fund", False)
            )
            budget_dict = {
                "id": budget.id,
                "project_id": budget.project_id,
                "category": budget.category or "Unknown",
                "amount": float(budget.amount),
                "period_type": budget.period_type,
                "start_date": budget.start_date.isoformat() if budget.start_date else None,
                "end_date": budget.end_date.isoformat() if budget.end_date else None,
                "is_active": budget.is_active,
                "spent_amount": spent,
                "remaining_amount": float(budget.amount) - spent,
                "spent_percentage": (spent / float(budget.amount) * 100) if budget.amount else 0,
            }
            budgets_list.append(budget_dict)
    
    # Process fund
    fund_data = None
    if not isinstance(fund_result, Exception):
        fund = fund_result.scalar_one_or_none()
        if fund:
            # Get fund transactions (already in transactions_list with from_fund=True)
            fund_transactions = [tx for tx in transactions_list if tx.get('from_fund', False)]
            total_deductions = sum(tx['amount'] for tx in fund_transactions if tx['type'] == 'Expense')
            total_additions_from_transactions = sum(tx['amount'] for tx in fund_transactions if tx['type'] == 'Income')
            
            # Calculate monthly additions if monthly_amount > 0
            total_monthly_additions = 0.0
            if fund.monthly_amount > 0:
                today = date.today()
                project = await ProjectRepository(db).get_by_id(project_id)
                if project and project.start_date:
                    calculation_start_date = project.start_date
                else:
                    calculation_start_date = fund.created_at.date() if hasattr(fund.created_at, 'date') else today
                total_monthly_additions = calculate_monthly_income_amount(
                    float(fund.monthly_amount),
                    calculation_start_date,
                    today
                )
            
            # Calculate initial_total: total amount that entered the fund from the beginning
            # Formula: initial_total = initial_balance + total_additions
            # We work backwards: initial_total = current_balance + total_deductions - total_additions
            stored_current_balance = float(fund.current_balance)
            calculated_initial = stored_current_balance + total_deductions - total_monthly_additions - total_additions_from_transactions
            # initial_total should represent actual money that entered, not debt
            # If negative, it means the fund started with a negative balance (debt), so no money actually entered
            initial_total = max(0, calculated_initial)
            
            # Recalculate current_balance from transactions to ensure it's correct
            # Formula: current_balance = initial_total + total_additions - total_deductions
            # But if initial_total is 0 (no money entered), then current_balance = total_additions - total_deductions
            total_additions = total_monthly_additions + total_additions_from_transactions
            recalculated_current_balance = initial_total + total_additions - total_deductions
            
            # Use the recalculated balance to ensure consistency
            current_balance = recalculated_current_balance
            
            fund_data = {
                "id": fund.id,
                "project_id": fund.project_id,
                "current_balance": current_balance,
                "monthly_amount": float(fund.monthly_amount),
                "total_deductions": total_deductions,
                "initial_total": initial_total,  # Total amount that entered the fund
                "transactions": fund_transactions
            }
    
    # Build expense categories list
    expense_categories = [
        {"category": cat, "amount": amount, "color": f"#{hash(cat) % 0xFFFFFF:06x}"}
        for cat, amount in expense_by_category.items()
    ]
    
    # Build project dict
    project_dict = {
        "id": project.id,
        "name": project.name,
        "description": project.description,
        "start_date": project.start_date.isoformat() if project.start_date else None,
        "end_date": project.end_date.isoformat() if project.end_date else None,
        "contract_duration_months": project.contract_duration_months,
        "budget_monthly": project.budget_monthly,
        "budget_annual": project.budget_annual,
        "manager_id": project.manager_id,
        "relation_project": project.relation_project,
        "is_parent_project": project.is_parent_project,
        "show_in_quotes_tab": getattr(project, 'show_in_quotes_tab', False),
        "is_active": project.is_active,
        "num_residents": project.num_residents,
        "monthly_price_per_apartment": project.monthly_price_per_apartment,
        "address": project.address,
        "city": project.city,
        "image_url": project.image_url,
        "contract_file_url": project.contract_file_url,
        "is_active": project.is_active,
        "created_at": project.created_at.isoformat() if project.created_at else None,
        "total_value": getattr(project, 'total_value', 0.0),
        "has_fund": fund_data is not None,
        "monthly_fund_amount": fund_data['monthly_amount'] if fund_data else None
    }
    
    # Build selected_period info if viewing a historical period
    selected_period_info = None
    if selected_period:
        # Hebrew letters for period labeling
        hebrew_letters = ['א', 'ב', 'ג', 'ד', 'ה', 'ו', 'ז', 'ח', 'ט', 'י']
        
        # Count periods in the same year for labeling
        periods_in_year = [
            p for year_data in periods_by_year_list 
            for p in year_data['periods'] 
            if year_data['year'] == selected_period.contract_year
        ]
        show_period_label = len(periods_in_year) > 1
        
        # Determine year_label
        if show_period_label:
            idx = selected_period.year_index - 1
            letter = hebrew_letters[idx] if idx < len(hebrew_letters) else str(idx + 1)
            year_label = f"תקופה {letter}"
        else:
            year_label = ""
        
        # Calculate financials for selected period
        contract_service = ContractPeriodService(db)
        period_financials = await contract_service._get_period_financials(selected_period)
        
        selected_period_info = {
            'period_id': selected_period.id,
            'start_date': selected_period.start_date.isoformat() if selected_period.start_date else None,
            'end_date': selected_period.end_date.isoformat() if selected_period.end_date else None,
            'contract_year': selected_period.contract_year,
            'year_index': selected_period.year_index,
            'year_label': year_label,
            'total_income': period_financials['total_income'],
            'total_expense': period_financials['total_expense'],
            'total_profit': period_financials['total_profit']
        }
    
    # Accepted price quote that was converted to this project (if any)
    accepted_quote = None
    quote_repo = QuoteProjectRepository(db)
    qp = await quote_repo.get_by_converted_project_id(project_id)
    if qp:
        accepted_quote = {
            "id": qp.id,
            "name": qp.name,
            "status": qp.status,
        }
    
    return {
        "project": project_dict,
        "transactions": transactions_list,
        "budgets": budgets_list,
        "expense_categories": expense_categories,
        "fund": fund_data,
        "current_period": current_period,
        "selected_period": selected_period_info,  # New field for historical period viewing
        "contract_periods": {
            "project_id": project_id,
            "periods_by_year": periods_by_year_list
        },
        "accepted_quote": accepted_quote,
    }


@router.get("/{project_id}/subprojects", response_model=list[ProjectOut])
async def get_subprojects(project_id: int, db: DBSessionDep, user = Depends(get_current_user)):
    """Get subprojects - accessible to all authenticated users"""
    from backend.services.fund_service import FundService
    
    subprojects = await ProjectRepository(db).get_subprojects(project_id)
    fund_service = FundService(db)
    
    # Add fund information to each subproject
    result = []
    for subproject in subprojects:
        fund = await fund_service.get_fund_by_project(subproject.id)
        subproject_dict = {
            "id": subproject.id,
            "name": subproject.name,
            "description": subproject.description,
            "start_date": subproject.start_date.isoformat() if subproject.start_date else None,
            "end_date": subproject.end_date.isoformat() if subproject.end_date else None,
            "contract_duration_months": subproject.contract_duration_months,
            "budget_monthly": subproject.budget_monthly,
            "budget_annual": subproject.budget_annual,
            "manager_id": subproject.manager_id,
            "relation_project": subproject.relation_project,
            "is_parent_project": subproject.is_parent_project,
            "num_residents": subproject.num_residents,
            "monthly_price_per_apartment": subproject.monthly_price_per_apartment,
            "address": subproject.address,
            "city": subproject.city,
            "image_url": subproject.image_url,
            "contract_file_url": subproject.contract_file_url,
            "is_active": subproject.is_active,
            "show_in_quotes_tab": getattr(subproject, 'show_in_quotes_tab', False),
            "created_at": subproject.created_at,
            "total_value": getattr(subproject, 'total_value', 0.0),
            "has_fund": fund is not None,
            "monthly_fund_amount": float(fund.monthly_amount) if fund else None
        }
        result.append(subproject_dict)
    
    return result

@router.get("/get_values/{project_id}", response_model=ProjectOut)
async def get_project_values(project_id: int, db: DBSessionDep, user = Depends(get_current_user)):
    """Get project values - accessible to all authenticated users"""
    project_data = await ProjectService(db).get_value_of_projects(project_id=project_id)
    if not project_data:
        raise HTTPException(status_code=404, detail="Project not found")
    return project_data

@router.post("/", response_model=ProjectOut)
async def create_project(db: DBSessionDep, data: ProjectCreate, user = Depends(require_permission("write", "project", project_id_param=None))):
    """Create project"""
    # Extract recurring transactions, budgets, and fund data from project data
    project_data = data.model_dump(exclude={'recurring_transactions', 'budgets', 'has_fund', 'monthly_fund_amount'})
    recurring_transactions = data.recurring_transactions or []
    budgets = data.budgets or []
    has_fund = data.has_fund or False
    monthly_fund_amount = data.monthly_fund_amount
    
    # Determine if this is a parent project or regular project
    # If relation_project is set, this is a subproject (not a parent project)
    # If is_parent_project is explicitly set to True, this is a parent project
    # Otherwise, if relation_project is None and is_parent_project is not explicitly False, default to regular project
    if project_data.get('relation_project') is not None:
        # This is a subproject - cannot be a parent project
        project_data['is_parent_project'] = False
        
        # Validate that the parent project exists and is actually a parent project
        parent_id = project_data['relation_project']
        parent_project = await ProjectRepository(db).get_by_id(parent_id)
        if not parent_project:
            raise HTTPException(status_code=404, detail=f"פרויקט אב עם מזהה {parent_id} לא נמצא")
        if not parent_project.is_parent_project:
            raise HTTPException(status_code=400, detail="לא ניתן ליצור תת-פרויקט לפרויקט רגיל. רק פרויקט על יכול לקבל תת-פרויקטים")
    else:
        # This is not a subproject - check if it should be a parent project
        # If is_parent_project is explicitly set, use that value
        # Otherwise, default to False (regular project)
        if 'is_parent_project' not in project_data:
            project_data['is_parent_project'] = False
    
    # פרויקט על: בלי תאריכים ומשך חוזה – רק שם ותיאור
    if project_data.get('is_parent_project') is True:
        project_data['start_date'] = None
        project_data['end_date'] = None
        project_data['contract_duration_months'] = None

    # ולידציה: מניעת שני פרויקטים (כל סוג) עם אותו שם
    name_trimmed = (project_data.get('name') or '').strip()
    if name_trimmed:
        existing = await ProjectRepository(db).get_project_by_name(name_trimmed)
        if existing:
            raise HTTPException(
                status_code=400,
                detail="קיים כבר פרויקט עם שם זה. יש לבחור שם ייחודי.",
            )
        project_data['name'] = name_trimmed

    # Create the project
    project = await ProjectService(db).create(user_id=user.id, **project_data)
    
    # Note: Fund creation is now handled by the frontend after project creation
    # The has_fund flag is still set on the project, but the actual fund creation
    # happens through a separate modal that allows the user to choose fund setup options
    
    # Create recurring transactions if provided
    if recurring_transactions:
        recurring_service = RecurringTransactionService(db)
        for rt_data in recurring_transactions:
            # Convert to dict and set the project_id for each recurring transaction
            rt_dict = rt_data.model_dump()
            rt_dict['project_id'] = project.id
            # Create new instance with project_id set
            rt_create = RecurringTransactionTemplateCreate(**rt_dict)
            await recurring_service.create_template(rt_create)
    
    # Create budgets if provided
    if budgets:
        budget_service = BudgetService(db)
        for idx, budget_data in enumerate(budgets):
            try:
                # Convert string dates to date objects
                from datetime import date as date_type
                start_date = None
                end_date = None

                if budget_data.start_date:
                    if isinstance(budget_data.start_date, str):
                        start_date = date_type.fromisoformat(budget_data.start_date)
                    else:
                        start_date = budget_data.start_date

                if budget_data.end_date:
                    if isinstance(budget_data.end_date, str):
                        end_date = date_type.fromisoformat(budget_data.end_date)
                    else:
                        end_date = budget_data.end_date

                logger.info(
                    "Creating budget for project %s: index=%d, category_id=%s, amount=%s, period_type=%s, start_date=%s, end_date=%s",
                    project.id, idx, budget_data.category_id, budget_data.amount, budget_data.period_type, start_date, end_date,
                )

                created_budget = await budget_service.create_budget(
                    project_id=project.id,
                    category_id=budget_data.category_id,
                    amount=budget_data.amount,
                    period_type=budget_data.period_type or "Annual",
                    start_date=start_date,
                    end_date=end_date
                )
                logger.info(
                    "Budget created: budget_id=%s, category_id=%s",
                    created_budget.id, budget_data.category_id,
                )
            except Exception as e:
                logger.exception(
                    "Failed to create budget for project %s: index=%d, category_id=%s, amount=%s, period_type=%s",
                    project.id, idx, budget_data.category_id, budget_data.amount, budget_data.period_type,
                )
                # Don't fail the entire project creation
    
    # Log create action with full details
    await AuditService(db).log_project_action(
        user_id=user.id,
        action='create',
        project_id=project.id,
        details={
            'name': project.name,
            'description': project.description,
            'budget_monthly': str(project.budget_monthly) if project.budget_monthly else None,
            'budget_annual': str(project.budget_annual) if project.budget_annual else None,
            'address': project.address,
            'city': project.city,
            'start_date': str(project.start_date) if project.start_date else None,
            'end_date': str(project.end_date) if project.end_date else None
        }
    )
    
    return project

@router.post("", response_model=ProjectOut)
async def create_project_no_slash(db: DBSessionDep, data: ProjectCreate, user = Depends(require_permission("write", "project", project_id_param=None))):
    """Alias without trailing slash to avoid 404 when redirect_slashes=False"""
    return await create_project(db, data, user)

@router.put("/{project_id}", response_model=ProjectOut)
async def update_project(project_id: int, db: DBSessionDep, data: ProjectUpdate, user = Depends(require_permission("update", "project", resource_id_param="project_id", project_id_param=None))):
    """Update project"""
    repo = ProjectRepository(db)
    project = await repo.get_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    budgets_to_add = data.budgets or []
    has_fund = data.has_fund
    monthly_fund_amount = data.monthly_fund_amount
    apply_from_period_id = data.apply_from_period_id

    # Store old values for audit log
    old_values = {
        'name': project.name,
        'description': project.description or '',
        'budget_monthly': str(project.budget_monthly) if project.budget_monthly else None,
        'budget_annual': str(project.budget_annual) if project.budget_annual else None,
        'address': project.address or '',
        'city': project.city or ''
    }

    update_payload = data.model_dump(exclude_unset=True, exclude={'budgets', 'has_fund', 'monthly_fund_amount', 'apply_from_period_id'})
    
    # Validate relation_project if it's being set
    if 'relation_project' in update_payload and update_payload['relation_project'] is not None:
        parent_id = update_payload['relation_project']
        parent_project = await repo.get_by_id(parent_id)
        if not parent_project:
            raise HTTPException(status_code=404, detail=f"פרויקט אב עם מזהה {parent_id} לא נמצא")
        if not parent_project.is_parent_project:
            raise HTTPException(status_code=400, detail="לא ניתן להגדיר פרויקט רגיל כפרויקט אב. רק פרויקט על יכול לקבל תת-פרויקטים")
        # If setting relation_project, this becomes a subproject (not a parent project)
        update_payload['is_parent_project'] = False
    elif 'relation_project' in update_payload and update_payload['relation_project'] is None:
        # If removing relation_project, we need to determine if it should be a parent project
        # Don't change is_parent_project if it's not explicitly set
        if 'is_parent_project' not in update_payload:
            # Keep current value
            pass
    
    # Prevent changing is_parent_project if project already has subprojects
    if 'is_parent_project' in update_payload:
        # Check if this project has subprojects
        subprojects = await repo.get_subprojects(project_id)
        if len(subprojects) > 0 and not update_payload['is_parent_project']:
            raise HTTPException(status_code=400, detail="לא ניתן לשנות פרויקט על לפרויקט רגיל כאשר יש לו תת-פרויקטים")
    
    # When updating name: ensure no duplicate project names (any project type)
    if 'name' in update_payload:
        name_trimmed = (update_payload['name'] or '').strip()
        if name_trimmed:
            existing = await repo.get_project_by_name(name_trimmed, exclude_project_id=project_id)
            if existing:
                raise HTTPException(
                    status_code=400,
                    detail="קיים כבר פרויקט עם שם זה. יש לבחור שם ייחודי.",
                )
            update_payload['name'] = name_trimmed

    # Handle contract_duration_months changes
    # If apply_from_period_id is provided, allow changing duration from that period onwards
    # Otherwise, prevent retroactive changes if there are past periods
    if 'contract_duration_months' in update_payload:
        new_duration = update_payload['contract_duration_months']
        old_duration = project.contract_duration_months
        
        # Only check if duration is actually changing
        if new_duration is not None and new_duration != old_duration:
            from backend.repositories.contract_period_repository import ContractPeriodRepository
            from sqlalchemy import select, or_
            from backend.services.contract_period_service import ContractPeriodService
            from dateutil.relativedelta import relativedelta
            
            period_repo = ContractPeriodRepository(db)
            today = date.today()
            
            # If apply_from_period_id is provided, regenerate periods from that period onwards
            if apply_from_period_id is not None:
                # Validate that the period exists and belongs to this project
                from_period = await period_repo.get_by_id(apply_from_period_id)
                if not from_period or from_period.project_id != project_id:
                    raise HTTPException(
                        status_code=404,
                        detail=f"תקופת חוזה עם מזהה {apply_from_period_id} לא נמצאה"
                    )
                
                # Validate new duration
                if new_duration <= 0:
                    raise HTTPException(
                        status_code=400,
                        detail="משך החוזה חייב להיות גדול מ-0"
                    )
                
                # Get all periods for this project
                all_periods = await period_repo.get_by_project(project_id)
                all_periods.sort(key=lambda p: p.start_date)
                
                # Find the period to start from
                from_period_idx = next((i for i, p in enumerate(all_periods) if p.id == apply_from_period_id), None)
                if from_period_idx is None:
                    raise HTTPException(
                        status_code=400,
                        detail="לא ניתן למצוא את התקופה שנבחרה"
                    )
                
                # Delete all periods from the selected period onwards (except the selected period itself)
                # We'll regenerate them with the new duration
                periods_to_delete = all_periods[from_period_idx + 1:]
                
                # Update the selected period's end_date based on new duration
                from_period.end_date = from_period.start_date + relativedelta(months=new_duration)
                from_period.contract_year = from_period.start_date.year
                await period_repo.update(from_period)
                
                # Delete future periods
                for period in periods_to_delete:
                    await period_repo.delete(period)
                
                # Regenerate periods from the selected period's end_date onwards
                contract_period_service = ContractPeriodService(db)
                current_start = from_period.end_date
                current_end = None
                
                # Generate periods until we reach today or beyond
                max_iterations = 200
                for i in range(max_iterations):
                    current_end = current_start + relativedelta(months=new_duration)
                    
                    # Stop if we've reached today or beyond
                    if current_end > today:
                        # Update project dates to match the last period
                        project.start_date = from_period.start_date
                        project.end_date = current_end
                        break
                    
                    # Create new period
                    contract_year = current_start.year
                    periods_in_year = await period_repo.get_by_project_and_year(project_id, contract_year)
                    year_index = len(periods_in_year) + 1
                    
                    new_period = ContractPeriod(
                        project_id=project_id,
                        start_date=current_start,
                        end_date=current_end,
                        contract_year=contract_year,
                        year_index=year_index
                    )
                    new_period = await period_repo.create(new_period)
                    
                    # Copy budgets from previous period
                    from backend.services.budget_service import BudgetService
                    budget_service = BudgetService(db)
                    previous_period_id = from_period.id if current_start == from_period.end_date else None
                    if previous_period_id is None:
                        # Find the previous period
                        prev_periods = await period_repo.get_by_project(project_id)
                        prev_periods.sort(key=lambda p: p.start_date)
                        for p in reversed(prev_periods):
                            if p.end_date == current_start:
                                previous_period_id = p.id
                                break
                    
                    await budget_service.copy_budgets_to_new_period(
                        project_id=project_id,
                        from_period_id=previous_period_id,
                        to_period=new_period,
                    )
                    
                    # Archive if period has ended
                    if current_end <= today:
                        from datetime import timezone as _tz
                        summary = await contract_period_service._get_period_financials(new_period)
                        new_period.is_archived = True
                        new_period.archived_at = datetime.now(_tz.utc).replace(tzinfo=None)
                        new_period.archived_by_user_id = user.id
                        new_period.total_income = summary['total_income']
                        new_period.total_expense = summary['total_expense']
                        new_period.total_profit = summary['total_profit']
                        await db.commit()
                    
                    current_start = current_end
                
                # Update project dates
                project.start_date = from_period.start_date
                if current_end and current_end > today:
                    project.end_date = current_end
                else:
                    # Find the last period
                    all_periods_after = await period_repo.get_by_project(project_id)
                    if all_periods_after:
                        last_period = max(all_periods_after, key=lambda p: p.end_date)
                        project.end_date = last_period.end_date
                
                await db.commit()
            else:
                # Old behavior: prevent retroactive changes if there are past periods
                periods = await period_repo.get_by_project(project_id)
                past_periods = [p for p in periods if p.end_date and p.end_date <= today]
                
                if len(past_periods) > 0:
                    raise HTTPException(
                        status_code=400, 
                        detail="לא ניתן לשנות את משך החוזה בחודשים לפרויקט שיש לו תקופות חוזה בעבר. יש לבחור תקופה להתחלה או לשנות את משך החוזה רק לפרויקטים חדשים."
                    )
    
    # NOTE: No end_date subtraction logic here.
    # Users expect the updated end_date to be exactly what they entered.
    
    updated_project = await ProjectService(db).update(project, **update_payload)
    
    # Handle fund creation/update
    fund_service = FundService(db)
    existing_fund = await fund_service.get_fund_by_project(project_id)
    
    if has_fund is not None:
        if has_fund:
            # Create or update fund (even if monthly_amount is 0)
            monthly_amount = monthly_fund_amount if monthly_fund_amount is not None and monthly_fund_amount > 0 else 0
            if existing_fund:
                # Update existing fund (repository already commits)
                await fund_service.update_fund(existing_fund, monthly_amount=monthly_amount)
            else:
                # Create new fund (repository already commits)
                # Calculate initial balance based on project start_date if it's in the past
                initial_balance = 0.0
                last_monthly_addition = None
                if monthly_amount > 0 and updated_project.start_date:
                    today = date.today()
                    if updated_project.start_date <= today:
                        # Calculate accumulated amount from contract start date to today
                        initial_balance = calculate_monthly_income_amount(
                            monthly_amount,
                            updated_project.start_date,
                            today
                        )
                        # Set last_monthly_addition to today to indicate all months up to today are accounted for
                        if initial_balance > 0:
                            last_monthly_addition = today
                
                await fund_service.create_fund(
                    project_id=project_id,
                    monthly_amount=monthly_amount,
                    initial_balance=initial_balance,
                    last_monthly_addition=last_monthly_addition
                )
        elif not has_fund and existing_fund:
            # Delete fund if has_fund is False (repository already commits)
            await fund_service.funds.delete(existing_fund)
    elif monthly_fund_amount is not None and existing_fund:
        # Update monthly amount only (if has_fund wasn't explicitly set) (repository already commits)
        monthly_amount = monthly_fund_amount if monthly_fund_amount > 0 else 0
        await fund_service.update_fund(existing_fund, monthly_amount=monthly_amount)

    # Handle new category budgets if provided
    if budgets_to_add:
        budget_service = BudgetService(db)
        for idx, budget_data in enumerate(budgets_to_add):
            try:
                from datetime import date as date_type
                start_date = None
                end_date = None

                if budget_data.start_date:
                    if isinstance(budget_data.start_date, str):
                        start_date = date_type.fromisoformat(budget_data.start_date)
                    else:
                        start_date = budget_data.start_date

                if budget_data.end_date:
                    if isinstance(budget_data.end_date, str):
                        end_date = date_type.fromisoformat(budget_data.end_date)
                    else:
                        end_date = budget_data.end_date

                logger.info(
                    "Adding budget during update for project %s: index=%d, category_id=%s, amount=%s, period_type=%s, start_date=%s, end_date=%s",
                    project_id, idx, budget_data.category_id, budget_data.amount, budget_data.period_type, start_date, end_date,
                )

                created_budget = await budget_service.create_budget(
                    project_id=project_id,
                    category_id=budget_data.category_id,
                    amount=budget_data.amount,
                    period_type=budget_data.period_type or "Annual",
                    start_date=start_date,
                    end_date=end_date
                )
                logger.info(
                    "Budget added during update: budget_id=%s, category_id=%s",
                    created_budget.id, budget_data.category_id,
                )
            except Exception as e:
                logger.exception(
                    "Failed to add budget during update for project %s: index=%d, category_id=%s, amount=%s, period_type=%s",
                    project_id, idx, budget_data.category_id, budget_data.amount, budget_data.period_type,
                )

    # Log update action with full details
    update_data = {k: str(v) for k, v in update_payload.items()}
    await AuditService(db).log_project_action(
        user_id=user.id,
        action='update',
        project_id=project_id,
        details={
            'project_name': project.name,
            'old_values': old_values,
            'new_values': update_data
        }
    )

    # Refresh project to get updated data including fund info
    await db.refresh(updated_project)
    
    # Get updated fund information for response
    fund = await fund_service.get_fund_by_project(project_id)
    
    # Convert to dict to modify
    # Explicitly convert dates to ISO format strings to avoid timezone issues
    project_dict = {
        "id": updated_project.id,
        "name": updated_project.name,
        "description": updated_project.description,
        "start_date": updated_project.start_date.isoformat() if updated_project.start_date else None,
        "end_date": updated_project.end_date.isoformat() if updated_project.end_date else None,
        "contract_duration_months": updated_project.contract_duration_months,
        "budget_monthly": updated_project.budget_monthly,
        "budget_annual": updated_project.budget_annual,
        "manager_id": updated_project.manager_id,
        "relation_project": updated_project.relation_project,
        "num_residents": updated_project.num_residents,
        "monthly_price_per_apartment": updated_project.monthly_price_per_apartment,
        "address": updated_project.address,
        "city": updated_project.city,
        "image_url": updated_project.image_url,
        "contract_file_url": updated_project.contract_file_url,
        "is_active": updated_project.is_active,
        "created_at": updated_project.created_at,
        "total_value": getattr(updated_project, 'total_value', 0.0),  # Use getattr with default value
        "has_fund": fund is not None,
        "monthly_fund_amount": float(fund.monthly_amount) if fund else None
    }
    
    return project_dict


@router.post("/{project_id}/upload-image", response_model=ProjectOut)
async def upload_project_image(project_id: int, db: DBSessionDep, file: UploadFile = File(...), user = Depends(require_permission("update", "project", resource_id_param="project_id", project_id_param=None))):
    """Upload project image to S3"""
    repo = ProjectRepository(db)
    project = await repo.get_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Validate file type (only images)
    allowed_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail=f"Invalid file type. Allowed types: {', '.join(allowed_extensions)}")

    # Upload to S3
    s3 = S3Service()
    content = await file.read()

    from io import BytesIO
    file_obj = BytesIO(content)

    image_url = s3.upload_file(
        prefix="projects",
        file_obj=file_obj,
        filename=file.filename or "project-image",
        content_type=file.content_type,
    )

    # Store full URL in image_url
    project.image_url = image_url
    await repo.update(project)

    return project


@router.post("/{project_id}/upload-contract", response_model=ProjectOut)
async def upload_project_contract(project_id: int, db: DBSessionDep, file: UploadFile = File(...), user = Depends(require_permission("update", "project", resource_id_param="project_id", project_id_param=None))):
    """Upload a building contract file for a project."""
    repo = ProjectRepository(db)
    project = await repo.get_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    allowed_extensions = {'.pdf', '.doc', '.docx', '.jpg', '.jpeg', '.png'}
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in allowed_extensions:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type. Allowed types: {', '.join(sorted(allowed_extensions))}"
        )

    s3 = S3Service()
    content = await file.read()

    from io import BytesIO
    file_obj = BytesIO(content)

    contract_url = s3.upload_file(
        prefix="project-contracts",
        file_obj=file_obj,
        filename=file.filename or "project-contract",
        content_type=file.content_type,
    )

    project.contract_file_url = contract_url
    await repo.update(project)

    return project


@router.post("/{project_id}/documents", response_model=dict)
async def upload_project_document(
    project_id: int, 
    db: DBSessionDep, 
    file: UploadFile = File(...),
    description: str | None = Form(None),
    user = Depends(require_permission("update", "project", resource_id_param="project_id", project_id_param=None))
):
    """Upload a document for a project"""
    repo = ProjectRepository(db)
    project = await repo.get_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    s3 = S3Service()
    content = await file.read()

    from io import BytesIO
    file_obj = BytesIO(content)

    file_url = s3.upload_file(
        prefix=f"projects/{project_id}/documents",
        file_obj=file_obj,
        filename=file.filename or "document",
        content_type=file.content_type,
    )

    from backend.models.document import Document
    from backend.repositories.document_repository import DocumentRepository
    doc = Document(entity_type="project", entity_id=project_id, file_path=file_url, description=description)
    doc = await DocumentRepository(db).create(doc)

    return {
        "id": doc.id,
        "file_path": doc.file_path,
        "description": doc.description,
        "uploaded_at": doc.uploaded_at.isoformat()
    }


@router.get("/{project_id}/documents", response_model=list[dict])
async def get_project_documents(project_id: int, db: DBSessionDep, user = Depends(get_current_user)):
    """Get all documents for a project"""
    from backend.repositories.document_repository import DocumentRepository
    docs = await DocumentRepository(db).list_by_project(project_id)

    return [
        {
            "id": doc.id,
            "file_path": doc.file_path,
            "description": doc.description,
            "uploaded_at": doc.uploaded_at.isoformat()
        } for doc in docs
    ]


@router.delete("/{project_id}/documents/{doc_id}")
async def delete_project_document(project_id: int, doc_id: int, db: DBSessionDep, user = Depends(require_permission("update", "project", resource_id_param="project_id", project_id_param=None))):
    """Delete a document from a project"""
    import asyncio
    from backend.repositories.document_repository import DocumentRepository
    doc_repo = DocumentRepository(db)
    doc = await doc_repo.get_by_id(doc_id)
    if not doc or doc.entity_type != "project" or doc.entity_id != project_id:
        raise HTTPException(status_code=404, detail="Document not found")

    file_path = doc.file_path
    await doc_repo.delete(doc)

    # Try to delete from S3
    if file_path and ("s3" in file_path.lower() or "amazonaws.com" in file_path):
        try:
            s3 = S3Service()
            await asyncio.to_thread(s3.delete_file, file_path)
        except Exception as e:
            logger.warning("Failed to delete file from S3: %s", e)

    return {"ok": True}


@router.post("/{project_id}/archive", response_model=ProjectOut)
async def archive_project(project_id: int, db: DBSessionDep, user = Depends(require_permission("delete", "project", resource_id_param="project_id", project_id_param=None))):
    """Archive project"""
    repo = ProjectRepository(db)
    project = await repo.get_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    archived = await repo.archive(project)
    
    # Log archive action
    await AuditService(db).log_project_action(
        user_id=user.id,
        action='archive',
        project_id=project_id,
        details={'name': project.name}
    )
    
    return archived


@router.post("/{project_id}/restore", response_model=ProjectOut)
async def restore_project(project_id: int, db: DBSessionDep, user = Depends(require_permission("delete", "project", resource_id_param="project_id", project_id_param=None))):
    """Restore project"""
    repo = ProjectRepository(db)
    project = await repo.get_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    restored = await repo.restore(project)
    
    # Log restore action
    await AuditService(db).log_project_action(
        user_id=user.id,
        action='restore',
        project_id=project_id,
        details={'name': project.name}
    )
    
    return restored


class DeleteProjectRequest(BaseModel):
    password: str


@router.delete("/{project_id}")
async def hard_delete_project(
    project_id: int, 
    delete_request: DeleteProjectRequest,
    db: DBSessionDep, 
    user = Depends(require_permission("delete", "project", resource_id_param="project_id", project_id_param=None))
):
    """Hard delete project, requires password verification"""
    # Verify password
    user_repo = UserRepository(db)
    db_user = await user_repo.get_by_id(user.id)
    if not db_user or not db_user.password_hash:
        raise HTTPException(status_code=400, detail="User not found or uses OAuth login")
    
    if not verify_password(delete_request.password, db_user.password_hash):
        raise HTTPException(status_code=400, detail="סיסמה שגויה")
    
    proj_repo = ProjectRepository(db)
    project = await proj_repo.get_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Store project details for audit log
    project_details = {'name': project.name}
    
    # Get all transactions before deletion to delete their files
    tx_repo = TransactionRepository(db)
    transactions = await tx_repo.list_by_project(project_id)
    
    # Delete all transaction files from S3 (only if S3 is configured)
    if settings.AWS_S3_BUCKET:
        try:
            s3_service = S3Service()
            for tx in transactions:
                if tx.file_path:
                    try:
                        s3_service.delete_file(tx.file_path)
                    except Exception as e:
                        # Log error but continue deletion
                        logger.warning("Failed to delete transaction file %s from S3: %s", tx.file_path, e)
            
            # Delete project contract file if exists
            if project.contract_file_url:
                try:
                    s3_service.delete_file(project.contract_file_url)
                except Exception as e:
                    # Log error but continue deletion
                    logger.warning("Failed to delete project contract file %s from S3: %s", project.contract_file_url, e)
        except (ValueError, Exception) as e:
            # If S3Service initialization fails (e.g., S3 not configured), log but continue with database deletion
            logger.warning("S3 service unavailable, skipping file deletion: %s", e)
    
    # Delete all related records before deleting the project
    # Order matters due to foreign key constraints
    
    # 1. Delete contract periods (cascades; archived info is now stored on the period itself)
    await db.execute(delete(ContractPeriod).where(ContractPeriod.project_id == project_id))
    
    # 3. Delete budgets
    await db.execute(delete(Budget).where(Budget.project_id == project_id))
    
    # 4. Delete recurring transaction templates
    await db.execute(delete(RecurringTransactionTemplate).where(RecurringTransactionTemplate.project_id == project_id))
    
    # 5. Delete subprojects
    await db.execute(delete(Subproject).where(Subproject.project_id == project_id))
    
    # 6. Delete fund
    await db.execute(delete(Fund).where(Fund.project_id == project_id))
    
    # 7. Delete transactions
    await tx_repo.delete_by_project(project_id)
    
    # 8. Nullify quote_projects references (project_id and converted_project_id) so quotes remain but project can be deleted
    await db.execute(update(QuoteProject).where(QuoteProject.project_id == project_id).values(project_id=None))
    await db.execute(update(QuoteProject).where(QuoteProject.converted_project_id == project_id).values(converted_project_id=None))
    
    # Commit all deletions and nullifications
    await db.commit()
    
    # 9. Finally, delete the project itself
    await proj_repo.delete(project)
    
    # Log delete action
    await AuditService(db).log_project_action(
        user_id=user.id,
        action='delete',
        project_id=project_id,
        details=project_details
    )
    
    return {"ok": True}


@router.get("/{project_id}/financial-summary")
async def get_parent_project_financial_summary(
    project_id: int,
    db: DBSessionDep,
    start_date: Optional[date] = Query(None, description="Start date for filtering transactions"),
    end_date: Optional[date] = Query(None, description="End date for filtering transactions"),
    user = Depends(get_current_user)
):
    """Get consolidated financial summary for a parent project including all subprojects"""
    # Use async approach instead of sync
    from sqlalchemy import select, and_, func
    from backend.models.project import Project
    from backend.models.transaction import Transaction
    from backend.services.project_service import calculate_start_date
    from datetime import date as date_type
    
    # Get parent project
    parent_result = await db.execute(
        select(Project).where(
            Project.id == project_id,
            Project.is_active == True
        )
    )
    parent_project = parent_result.scalar_one_or_none()
    
    if not parent_project:
        raise HTTPException(status_code=404, detail="Parent project not found")
    
    # Get all subprojects
    subprojects_result = await db.execute(
        select(Project).where(
            Project.relation_project == project_id,
            Project.is_active == True
        )
    )
    subprojects = subprojects_result.scalars().all()
    
    # If no start_date provided, use project start_date (or 1 year back as fallback if no start_date)
    if not start_date:
        if parent_project.start_date:
            start_date = parent_project.start_date
        else:
            # Fallback: use 1 year ago if no project start date
            from dateutil.relativedelta import relativedelta
            start_date = date_type.today() - relativedelta(years=1)
    
    # If no end_date provided, use today
    if not end_date:
        end_date = date_type.today()
    
    # Build date filter
    date_conditions = []
    if start_date:
        date_conditions.append(Transaction.tx_date >= start_date)
    if end_date:
        date_conditions.append(Transaction.tx_date <= end_date)
    
    # Get transactions for parent project
    parent_transactions_query = select(Transaction).where(Transaction.project_id == project_id)
    if date_conditions:
        parent_transactions_query = parent_transactions_query.where(and_(*date_conditions))
    
    parent_transactions_result = await db.execute(parent_transactions_query)
    parent_transactions = parent_transactions_result.scalars().all()
    
    # Calculate parent project financials
    parent_transaction_income = sum(float(t.amount) for t in parent_transactions if t.type == 'Income' and not t.from_fund)
    parent_expense = sum(float(t.amount) for t in parent_transactions if t.type == 'Expense' and not t.from_fund)
    
    # Calculate income from parent project's monthly budget (treated as expected monthly income)
    parent_project_income = 0.0
    monthly_income = float(parent_project.budget_monthly or 0)
    if monthly_income > 0:
        parent_transaction_income = 0.0
        # Use project start_date if available, otherwise use created_at date
        if parent_project.start_date:
            income_calculation_start = parent_project.start_date
        elif hasattr(parent_project, 'created_at') and parent_project.created_at:
            try:
                if hasattr(parent_project.created_at, 'date'):
                    income_calculation_start = parent_project.created_at.date()
                elif isinstance(parent_project.created_at, date):
                    income_calculation_start = parent_project.created_at
                else:
                    # Fallback: use start_date parameter
                    income_calculation_start = start_date
            except (AttributeError, TypeError):
                # Fallback: use start_date parameter
                income_calculation_start = start_date
        else:
            # Fallback: use start_date parameter (which is already set to project start or 1 year ago)
            income_calculation_start = start_date
        parent_project_income = calculate_monthly_income_amount(monthly_income, income_calculation_start, end_date)
    
    parent_income = parent_transaction_income + parent_project_income
    parent_profit = parent_income - parent_expense
    parent_profit_margin = (parent_profit / parent_income * 100) if parent_income > 0 else 0
    
    # Calculate subproject financials
    subproject_financials = []
    total_subproject_income = 0
    total_subproject_expense = 0
    
    for subproject in subprojects:
        subproject_transactions_query = select(Transaction).where(Transaction.project_id == subproject.id)
        if date_conditions:
            subproject_transactions_query = subproject_transactions_query.where(and_(*date_conditions))
        
        subproject_transactions_result = await db.execute(subproject_transactions_query)
        subproject_transactions = subproject_transactions_result.scalars().all()
        
        subproject_transaction_income = sum(float(t.amount) for t in subproject_transactions if t.type == 'Income' and not t.from_fund)
        subproject_expense = sum(float(t.amount) for t in subproject_transactions if t.type == 'Expense' and not t.from_fund)
        
        # Calculate income from subproject monthly budget (treated as expected monthly income)
        subproject_project_income = 0.0
        subproject_monthly_income = float(subproject.budget_monthly or 0)
        if subproject_monthly_income > 0:
            subproject_transaction_income = 0.0
            # Use project start_date if available, otherwise use created_at date
            if subproject.start_date:
                income_calculation_start = subproject.start_date
            elif hasattr(subproject, 'created_at') and subproject.created_at:
                try:
                    if hasattr(subproject.created_at, 'date'):
                        income_calculation_start = subproject.created_at.date()
                    elif isinstance(subproject.created_at, date):
                        income_calculation_start = subproject.created_at
                    else:
                        # Fallback: use start_date parameter
                        income_calculation_start = start_date
                except (AttributeError, TypeError):
                    # Fallback: use start_date parameter
                    income_calculation_start = start_date
            else:
                # Fallback: use start_date parameter (which is already set to project start or 1 year ago)
                income_calculation_start = start_date
            subproject_project_income = calculate_monthly_income_amount(subproject_monthly_income, income_calculation_start, end_date)
        
        subproject_income = subproject_transaction_income + subproject_project_income
        subproject_profit = subproject_income - subproject_expense
        subproject_profit_margin = (subproject_profit / subproject_income * 100) if subproject_income > 0 else 0
        
        # Determine status
        if subproject_profit_margin >= 10:
            status = 'green'
        elif subproject_profit_margin <= -10:
            status = 'red'
        else:
            status = 'yellow'
        
        subproject_financials.append({
            'id': subproject.id,
            'name': subproject.name,
            'income': subproject_income,
            'expense': subproject_expense,
            'profit': subproject_profit,
            'profit_margin': subproject_profit_margin,
            'status': status
        })
        
        total_subproject_income += subproject_income
        total_subproject_expense += subproject_expense
    
    # Calculate consolidated totals
    total_income = parent_income + total_subproject_income
    total_expense = parent_expense + total_subproject_expense
    total_profit = total_income - total_expense
    total_profit_margin = (total_profit / total_income * 100) if total_income > 0 else 0
    
    return {
        'parent_project': {
            'id': parent_project.id,
            'name': parent_project.name,
            'description': parent_project.description,
            'address': parent_project.address,
            'city': parent_project.city,
            'num_residents': parent_project.num_residents,
            'monthly_price_per_apartment': parent_project.monthly_price_per_apartment,
            'budget_monthly': parent_project.budget_monthly,
            'budget_annual': parent_project.budget_annual
        },
        'financial_summary': {
            'total_income': total_income,
            'total_expense': total_expense,
            'net_profit': total_profit,
            'profit_margin': total_profit_margin,
            'subproject_count': len(subprojects),
            'active_subprojects': len([sp for sp in subprojects if sp.is_active])
        },
        'parent_financials': {
            'income': parent_income,
            'expense': parent_expense,
            'profit': parent_profit,
            'profit_margin': parent_profit_margin
        },
        'subproject_financials': subproject_financials,
        'date_range': {
            'start_date': start_date.isoformat() if start_date else None,
            'end_date': end_date.isoformat() if end_date else None
        }
    }


@router.get("/{project_id}/fund")
async def get_project_fund(
    project_id: int,
    db: DBSessionDep,
    user = Depends(get_current_user)
):
    """Get fund details for a project"""
    from backend.schemas.fund import FundWithTransactions
    from sqlalchemy import select, and_, func
    from backend.models.transaction import Transaction
    from datetime import date
    
    fund_service = FundService(db)
    fund = await fund_service.get_fund_by_project(project_id)
    
    if not fund:
        raise HTTPException(status_code=404, detail="Fund not found for this project")
    
    # Ensure monthly addition is made if needed
    await fund_service.ensure_monthly_addition(project_id)
    # Refresh fund after potential update
    fund = await fund_service.get_fund_by_project(project_id)
    
    # Get transactions from fund
    transactions_query = select(Transaction).where(
        and_(
            Transaction.project_id == project_id,
            Transaction.from_fund == True
        )
    ).order_by(Transaction.tx_date.desc())
    
    transactions_result = await db.execute(transactions_query)
    transactions = transactions_result.scalars().all()
    
    # Calculate total deductions (total amount withdrawn from fund - Expense transactions)
    total_deductions_query = select(func.coalesce(func.sum(Transaction.amount), 0)).where(
        and_(
            Transaction.project_id == project_id,
            Transaction.from_fund == True,
            Transaction.type == 'Expense'
        )
    )
    total_deductions_result = await db.execute(total_deductions_query)
    total_deductions = float(total_deductions_result.scalar_one())
    
    # Calculate total additions from Income transactions to fund
    total_additions_from_transactions_query = select(func.coalesce(func.sum(Transaction.amount), 0)).where(
        and_(
            Transaction.project_id == project_id,
            Transaction.from_fund == True,
            Transaction.type == 'Income'
        )
    )
    total_additions_from_transactions_result = await db.execute(total_additions_from_transactions_query)
    total_additions_from_transactions = float(total_additions_from_transactions_result.scalar_one())
    
    # Calculate initial balance and total monthly additions
    monthly_amount = float(fund.monthly_amount)
    
    # Get project to access start_date
    project_repo = ProjectRepository(db)
    project = await project_repo.get_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Calculate total monthly additions based on project start date (not fund creation date)
    # Use project.start_date if available, otherwise use fund.created_at date
    total_monthly_additions = 0.0
    if monthly_amount > 0:
        today = date.today()
        
        # Determine the calculation start date: prefer project.start_date, fallback to fund.created_at
        if project.start_date:
            calculation_start_date = project.start_date
        else:
            calculation_start_date = fund.created_at.date() if hasattr(fund.created_at, 'date') else date.today()
        
        # Use calculate_monthly_income_amount to get the correct total from start date to today
        # This function properly handles monthly occurrences from start_date to current_date
        total_monthly_additions = calculate_monthly_income_amount(
            monthly_amount, 
            calculation_start_date, 
            today
        )
    
    # Total additions = monthly additions + income transactions to fund
    total_additions = total_monthly_additions + total_additions_from_transactions
    
    # Recalculate current_balance from transactions (same logic as get_project_full) so it matches
    # the "פרטי הקופה" display. If the stored balance drifted (e.g. missing deduction), sync it.
    stored_current_balance = float(fund.current_balance)
    calculated_initial = stored_current_balance + total_deductions - total_monthly_additions - total_additions_from_transactions
    initial_total = max(0.0, calculated_initial)
    recalculated_current_balance = initial_total + total_additions - total_deductions
    if abs(recalculated_current_balance - stored_current_balance) > 0.01:
        await fund_service.update_fund(fund, current_balance=recalculated_current_balance)
        fund = await fund_service.get_fund_by_project(project_id)
    
    current_balance = float(fund.current_balance)
    initial_balance = current_balance - total_additions + total_deductions
    initial_total = current_balance + total_deductions

    # Load user repository for created_by_user
    from backend.repositories.user_repository import UserRepository
    from backend.repositories.document_repository import DocumentRepository
    user_repo = UserRepository(db)
    doc_repo = DocumentRepository(db)
    
    # Convert transactions to dict with additional info
    transactions_list = []
    for tx in transactions:
        # Get creator user info
        created_by_user = None
        if hasattr(tx, 'created_by_user_id') and tx.created_by_user_id:
            creator = await user_repo.get_by_id(tx.created_by_user_id)
            if creator:
                created_by_user = {
                    'id': creator.id,
                    'full_name': creator.full_name,
                    'email': creator.email
                }
        
        # Get documents count
        documents_count = 0
        try:
            documents = await doc_repo.get_by_transaction_id(tx.id)
            documents_count = len(documents) if documents else 0
        except Exception:
            logger.warning("Failed to get documents count for transaction %s", tx.id)
        
        transactions_list.append({
            'id': tx.id,
            'tx_date': tx.tx_date.isoformat() if tx.tx_date else None,
            'type': tx.type,
            'amount': float(tx.amount),
            'description': tx.description,
            'category': tx.category,
            'notes': tx.notes,
            'created_by_user': created_by_user,
            'file_path': getattr(tx, 'file_path', None),
            'documents_count': documents_count
        })
    
    return {
        'id': fund.id,
        'project_id': fund.project_id,
        'current_balance': float(fund.current_balance),
        'monthly_amount': monthly_amount,
        'last_monthly_addition': fund.last_monthly_addition.isoformat() if fund.last_monthly_addition else None,
        'created_at': fund.created_at.isoformat(),
        'updated_at': fund.updated_at.isoformat(),
        'initial_balance': initial_balance,
        'initial_total': initial_total,  # Initial balance + all monthly additions
        'total_additions': total_additions,  # Total monthly additions made
        'total_deductions': total_deductions,  # Total amount withdrawn from fund
        'transactions': transactions_list
    }


@router.post("/{project_id}/fund")
async def create_project_fund(
    db: DBSessionDep,
    project_id: int,
    monthly_amount: float = Query(0, description="Monthly amount to add to fund"),
    initial_balance: Optional[float] = Query(None, description="Initial balance for the fund"),
    last_monthly_addition: Optional[str] = Query(None, description="Last monthly addition date (YYYY-MM-DD)"),
    user = Depends(get_current_user)
):
    """Create a fund for an existing project"""
    # Check if project exists
    project_repo = ProjectRepository(db)
    project = await project_repo.get_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Check if fund already exists
    fund_service = FundService(db)
    existing_fund = await fund_service.get_fund_by_project(project_id)
    if existing_fund:
        raise HTTPException(status_code=400, detail="Fund already exists for this project")
    
    # Use provided initial_balance, or calculate based on project start_date if not provided
    final_initial_balance = 0.0
    final_last_monthly_addition = None
    
    if initial_balance is not None:
        # Use provided initial_balance
        final_initial_balance = initial_balance
        if last_monthly_addition:
            try:
                final_last_monthly_addition = date.fromisoformat(last_monthly_addition)
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid last_monthly_addition date format. Use YYYY-MM-DD")
    else:
        # Calculate initial balance based on project start_date if it's in the past (backward compatibility)
        if monthly_amount > 0 and project.start_date:
            today = date.today()
            if project.start_date <= today:
                # Calculate accumulated amount from contract start date to today
                final_initial_balance = calculate_monthly_income_amount(
                    monthly_amount,
                    project.start_date,
                    today
                )
                # Set last_monthly_addition to today to indicate all months up to today are accounted for
                if final_initial_balance > 0:
                    final_last_monthly_addition = today
    
    # Create fund
    fund = await fund_service.create_fund(
        project_id=project_id,
        monthly_amount=monthly_amount,
        initial_balance=final_initial_balance,
        last_monthly_addition=final_last_monthly_addition
    )
    
    return {
        'id': fund.id,
        'project_id': fund.project_id,
        'current_balance': float(fund.current_balance),
        'monthly_amount': float(fund.monthly_amount),
        'created_at': fund.created_at.isoformat()
    }


@router.put("/{project_id}/fund")
async def update_project_fund(
    db: DBSessionDep,
    project_id: int,
    monthly_amount: Optional[float] = Query(None, description="Monthly amount to add to fund"),
    current_balance: Optional[float] = Query(None, description="Current balance of the fund"),
    update_scope: Optional[str] = Query(None, description="Scope of update: from_start, from_this_month, only_this_month"),
    user = Depends(get_current_user)
):
    """Update fund monthly amount and/or balance for a project"""
    # Check if project exists
    project_repo = ProjectRepository(db)
    project = await project_repo.get_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Check if fund exists
    fund_service = FundService(db)
    fund = await fund_service.get_fund_by_project(project_id)
    if not fund:
        raise HTTPException(status_code=404, detail="Fund not found for this project")
    
    # Update fund
    update_data = {}
    if monthly_amount is not None:
        update_data['monthly_amount'] = monthly_amount
    if current_balance is not None:
        update_data['current_balance'] = current_balance
    
    await fund_service.update_fund(
        fund, 
        update_scope=update_scope, 
        project_start_date=project.start_date, 
        **update_data
    )
    
    return {
        'id': fund.id,
        'project_id': fund.project_id,
        'current_balance': float(fund.current_balance),
        'monthly_amount': float(fund.monthly_amount),
        'updated_at': fund.updated_at.isoformat()
    }


class DeleteFundRequest(BaseModel):
    password: str


@router.delete("/{project_id}/fund")
async def delete_project_fund(
    project_id: int,
    delete_request: DeleteFundRequest,
    db: DBSessionDep,
    user = Depends(get_current_user)
):
    """Delete fund for a project (and all fund transactions + documents), requires password verification"""
    user_repo = UserRepository(db)
    db_user = await user_repo.get_by_id(user.id)
    if not db_user or not db_user.password_hash:
        raise HTTPException(status_code=400, detail="User not found or uses OAuth login")

    if not verify_password(delete_request.password, db_user.password_hash):
        raise HTTPException(status_code=400, detail="סיסמה שגויה")

    fund_service = FundService(db)
    fund = await fund_service.get_fund_by_project(project_id)
    if not fund:
        raise HTTPException(status_code=404, detail="Fund not found")

    # Fetch all fund transactions before deletion (for S3 cleanup)
    from sqlalchemy import select, delete as sql_delete, and_
    from backend.models.transaction import Transaction
    from backend.models.document import Document

    fund_txs_result = await db.execute(
        select(Transaction).where(
            and_(Transaction.project_id == project_id, Transaction.from_fund == True)
        )
    )
    fund_transactions = list(fund_txs_result.scalars().all())
    fund_tx_ids = [tx.id for tx in fund_transactions]

    # Collect all S3 file paths to delete
    s3_paths: list[str] = []
    for tx in fund_transactions:
        if tx.file_path:
            s3_paths.append(tx.file_path)

    # Fetch document file paths for S3 deletion
    if fund_tx_ids:
        docs_result = await db.execute(
            select(Document).where(Document.transaction_id.in_(fund_tx_ids))
        )
        for doc in docs_result.scalars().all():
            if doc.file_path:
                s3_paths.append(doc.file_path)

    # Delete S3 files (only if S3 is configured)
    if settings.AWS_S3_BUCKET and s3_paths:
        try:
            s3_service = S3Service()
            for path in s3_paths:
                try:
                    s3_service.delete_file(path)
                except Exception as e:
                    logger.warning("Failed to delete file %s from S3: %s", path, e)
        except Exception as e:
            logger.warning("S3 service unavailable, skipping file deletion: %s", e)

    # Delete fund transactions from DB (documents cascade via FK ondelete=CASCADE)
    if fund_tx_ids:
        await db.execute(sql_delete(Transaction).where(Transaction.id.in_(fund_tx_ids)))

    # Delete the fund
    await db.delete(fund)
    await db.commit()

    return {"detail": "הקופה נמחקה בהצלחה"}


@router.get("/{project_id}/financial-trends")
async def get_financial_trends(
    project_id: int,
    db: DBSessionDep,
    years_back: int = Query(5, description="Number of years to look back"),
    user = Depends(get_current_user)
):
    """Get financial trends over the last N years"""
    from sqlalchemy import select, and_, func, extract
    from backend.models.project import Project
    from backend.models.transaction import Transaction
    from datetime import datetime, date
    
    # Get parent project
    parent_result = await db.execute(
        select(Project).where(
            Project.id == project_id,
            Project.is_active == True
        )
    )
    parent_project = parent_result.scalar_one_or_none()
    
    if not parent_project:
        raise HTTPException(status_code=404, detail="Parent project not found")
    
    # Get all subprojects
    subprojects_result = await db.execute(
        select(Project).where(
            Project.relation_project == project_id,
            Project.is_active == True
        )
    )
    subprojects = subprojects_result.scalars().all()
    
    # Calculate trends for the last N years
    trends = []
    current_year = datetime.now().year
    
    for i in range(years_back):
        year = current_year - i
        
        # Get start and end of year
        year_start = date(year, 1, 1)
        year_end = date(year, 12, 31)
        
        # Get transactions for parent project in this year
        parent_transactions_query = select(Transaction).where(
            and_(
                Transaction.project_id == project_id,
                Transaction.tx_date >= year_start,
                Transaction.tx_date <= year_end
            )
        )
        parent_transactions_result = await db.execute(parent_transactions_query)
        parent_transactions = parent_transactions_result.scalars().all()
        
        parent_income = sum(t.amount for t in parent_transactions if t.type == 'Income')
        parent_expense = sum(t.amount for t in parent_transactions if t.type == 'Expense')
        
        # Get transactions for subprojects in this year
        total_subproject_income = 0
        total_subproject_expense = 0
        
        for subproject in subprojects:
            subproject_transactions_query = select(Transaction).where(
                and_(
                    Transaction.project_id == subproject.id,
                    Transaction.tx_date >= year_start,
                    Transaction.tx_date <= year_end
                )
            )
            subproject_transactions_result = await db.execute(subproject_transactions_query)
            subproject_transactions = subproject_transactions_result.scalars().all()
            
            subproject_income = sum(t.amount for t in subproject_transactions if t.type == 'Income')
            subproject_expense = sum(t.amount for t in subproject_transactions if t.type == 'Expense')
            
            total_subproject_income += subproject_income
            total_subproject_expense += subproject_expense
        
        # Calculate totals
        total_income = parent_income + total_subproject_income
        total_expense = parent_expense + total_subproject_expense
        total_profit = total_income - total_expense
        total_profit_margin = (total_profit / total_income * 100) if total_income > 0 else 0
        
        trends.append({
            'year': year,
            'income': total_income,
            'expense': total_expense,
            'profit': total_profit,
            'profit_margin': total_profit_margin
        })
    
    # Reverse to get chronological order
    trends.reverse()
    
    return {
        'trends': trends,
        'period_years': years_back
    }


# ============================================================================
# Contract Period Endpoints
# ============================================================================

@router.get("/{project_id}/contract-periods/current")
async def get_current_contract_period(
    project_id: int,
    db: DBSessionDep,
    user = Depends(get_current_user)
):
    """Get the current active contract period for a project"""
    service = ContractPeriodService(db)
    current_period = await service.get_current_contract_period(project_id)
    
    if not current_period:
        return {
            'project_id': project_id,
            'current_period': None
        }
    
    return {
        'project_id': project_id,
        'current_period': current_period
    }

@router.get("/{project_id}/contract-periods")
async def get_previous_contract_periods(
    project_id: int,
    db: DBSessionDep,
    user = Depends(get_current_user)
):
    """Get all previous contract periods grouped by year for a project"""
    service = ContractPeriodService(db)
    periods_by_year = await service.get_previous_contracts_by_year(project_id)
    
    # Convert to list format for easier frontend handling
    result = []
    # Sort years in descending order (most recent first)
    for year in sorted(periods_by_year.keys(), reverse=True):
        result.append({
            'year': year,
            'periods': periods_by_year[year]
        })
    
    return {
        'project_id': project_id,
        'periods_by_year': result
    }


@router.get("/{project_id}/contract-periods/summary/by-dates")
async def get_contract_period_summary_by_dates(
    project_id: int,
    start_date: str,
    db: DBSessionDep,
    user = Depends(get_current_user),
    end_date: Optional[str] = None
):
    """Get full summary of a contract period by dates (for virtual periods)"""
    from datetime import date as date_type
    service = ContractPeriodService(db)
    
    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date) if end_date else None
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        
    summary = await service.get_contract_period_summary(
        project_id=project_id,
        start_date=start,
        end_date=end
    )
    
    if not summary:
        raise HTTPException(status_code=404, detail="Contract period summary could not be generated")
    
    return summary


@router.get("/{project_id}/contract-periods/{period_id}")
async def get_contract_period_summary(
    project_id: int,
    period_id: int,
    db: DBSessionDep,
    user = Depends(get_current_user)
):
    """Get full summary of a contract period including transactions and budgets (read-only)"""
    service = ContractPeriodService(db)
    summary = await service.get_contract_period_summary(period_id)
    
    if not summary:
        raise HTTPException(status_code=404, detail="Contract period not found")
    
    if summary['project_id'] != project_id:
        raise HTTPException(status_code=400, detail="Contract period does not belong to this project")
    
    return summary


@router.put("/{project_id}/contract-periods/{period_id}")
async def update_contract_period(
    project_id: int,
    period_id: int,
    start_date: Optional[str] = Body(None),
    end_date: Optional[str] = Body(None),
    db: DBSessionDep = None,
    user = Depends(require_permission("delete", "project", resource_id_param="project_id", project_id_param=None))
):
    """Update contract period dates"""
    service = ContractPeriodService(db)
    
    # Parse dates
    start_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
             raise HTTPException(status_code=400, detail="Invalid start_date format. Use YYYY-MM-DD")

    end_date_obj = None
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
             raise HTTPException(status_code=400, detail="Invalid end_date format. Use YYYY-MM-DD")

    updated = await service.update_period_dates(period_id, start_date_obj, end_date_obj)
    if not updated:
        raise HTTPException(status_code=404, detail="Contract period not found")
        
    return {"success": True}


@router.get("/{project_id}/contract-periods/{period_id}/export-csv")
async def export_contract_period_csv(
    project_id: int,
    period_id: int,
    db: DBSessionDep,
    user = Depends(get_current_user),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Export contract period data to Excel with colors and formatting"""
    try:
        service = ContractPeriodService(db)
        if period_id > 0:
            summary = await service.get_contract_period_summary(period_id)
        else:
            from datetime import date as date_type
            start = date_type.fromisoformat(start_date) if start_date else date_type.today()
            end = date_type.fromisoformat(end_date) if end_date else date_type.today()
            summary = await service.get_contract_period_summary(
                project_id=project_id,
                start_date=start,
                end_date=end
            )
        
        if not summary:
            raise HTTPException(status_code=404, detail="Contract period not found")
        
        if summary['project_id'] != project_id:
            raise HTTPException(status_code=400, detail="Contract period does not belong to this project")
        
        # Get project name
        project = await ProjectRepository(db).get_by_id(project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "סיכום תקופת חוזה"
            ws.sheet_view.rightToLeft = True
            ws.sheet_view.rightToLeft = True
            
            # Define colors and styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            income_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            expense_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            profit_positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            profit_negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            header_font = Font(bold=True, color="FFFFFF", size=12)
            title_font = Font(bold=True, color="FFFFFF", size=14)
            normal_font = Font(size=11)
            bold_font = Font(bold=True, size=11)
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            row = 1
            
            # Title
            ws.merge_cells(f'A{row}:B{row}')
            cell = ws[f'A{row}']
            cell.value = 'סיכום תקופת חוזה'
            cell.fill = title_fill
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 2
            
            # Project info
            ws[f'A{row}'] = 'שם פרויקט'
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = project.name
            row += 1
            
            ws[f'A{row}'] = 'שנת חוזה'
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = summary['year_label']
            row += 1
            
            ws[f'A{row}'] = 'תאריך התחלה'
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = summary['start_date']
            row += 1
            
            ws[f'A{row}'] = 'תאריך סיום'
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = summary['end_date']
            row += 2
            
            # Financial Summary
            ws.merge_cells(f'A{row}:B{row}')
            cell = ws[f'A{row}']
            cell.value = 'סיכום כלכלי'
            cell.fill = section_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            # Financial headers
            ws[f'A{row}'] = 'סוג'
            ws[f'B{row}'] = 'סכום (₪)'
            for col in ['A', 'B']:
                cell = ws[f'{col}{row}']
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            # Income
            ws[f'A{row}'] = 'סה"כ הכנסות'
            ws[f'B{row}'] = summary['total_income']
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'].fill = income_fill
            ws[f'B{row}'].number_format = '#,##0.00'
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = border
            row += 1
            
            # Expense
            ws[f'A{row}'] = 'סה"כ הוצאות'
            ws[f'B{row}'] = summary['total_expense']
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'].fill = expense_fill
            ws[f'B{row}'].number_format = '#,##0.00'
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = border
            row += 1
            
            # Profit
            ws[f'A{row}'] = 'סה"כ רווח'
            ws[f'B{row}'] = summary['total_profit']
            ws[f'A{row}'].font = bold_font
            profit_fill = profit_positive_fill if summary['total_profit'] >= 0 else profit_negative_fill
            ws[f'B{row}'].fill = profit_fill
            ws[f'B{row}'].number_format = '#,##0.00'
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = border
            row += 2
            
            # Check if this is a previous year period and add year summary if so
            from datetime import date
            current_year = date.today().year
            contract_year = summary.get('contract_year', current_year)
            
            # If this is a previous year (not current year), add summary of all periods in that year
            if contract_year < current_year:
                from backend.repositories.contract_period_repository import ContractPeriodRepository
                period_repo = ContractPeriodRepository(db)
                all_year_periods = await period_repo.get_by_project_and_year(project_id, contract_year)
                
                if len(all_year_periods) > 0:  # Show if there are any periods
                    # Calculate year totals
                    year_total_income = sum(float(p.total_income or 0) for p in all_year_periods)
                    year_total_expense = sum(float(p.total_expense or 0) for p in all_year_periods)
                    year_total_profit = year_total_income - year_total_expense
                    
                    # Year summary section header
                    ws.merge_cells(f'A{row}:B{row}')
                    cell = ws[f'A{row}']
                    cell.value = f'סיכום כולל של כל התקופות בשנה {contract_year}'
                    cell.fill = section_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    row += 1
                    
                    ws[f'A{row}'] = 'מספר תקופות'
                    ws[f'A{row}'].font = bold_font
                    ws[f'B{row}'] = len(all_year_periods)
                    row += 2
                    
                    # Year financial summary
                    ws.merge_cells(f'A{row}:B{row}')
                    cell = ws[f'A{row}']
                    cell.value = 'סיכום כלכלי כולל של השנה'
                    cell.fill = section_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    row += 1
                    
                    # Year financial headers
                    ws[f'A{row}'] = 'סוג'
                    ws[f'B{row}'] = 'סכום (₪)'
                    for col in ['A', 'B']:
                        cell = ws[f'{col}{row}']
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    row += 1
                    
                    # Year Income
                    ws[f'A{row}'] = 'סה"כ הכנסות'
                    ws[f'B{row}'] = year_total_income
                    ws[f'A{row}'].font = bold_font
                    ws[f'B{row}'].fill = income_fill
                    ws[f'B{row}'].number_format = '#,##0.00'
                    for col in ['A', 'B']:
                        ws[f'{col}{row}'].border = border
                    row += 1
                    
                    # Year Expense
                    ws[f'A{row}'] = 'סה"כ הוצאות'
                    ws[f'B{row}'] = year_total_expense
                    ws[f'A{row}'].font = bold_font
                    ws[f'B{row}'].fill = expense_fill
                    ws[f'B{row}'].number_format = '#,##0.00'
                    for col in ['A', 'B']:
                        ws[f'{col}{row}'].border = border
                    row += 1
                    
                    # Year Profit
                    ws[f'A{row}'] = 'סה"כ רווח'
                    ws[f'B{row}'] = year_total_profit
                    ws[f'A{row}'].font = bold_font
                    year_profit_fill = profit_positive_fill if year_total_profit >= 0 else profit_negative_fill
                    ws[f'B{row}'].fill = year_profit_fill
                    ws[f'B{row}'].number_format = '#,##0.00'
                    for col in ['A', 'B']:
                        ws[f'{col}{row}'].border = border
                    row += 2
                    
                    # Breakdown by period
                    ws.merge_cells(f'A{row}:F{row}')
                    cell = ws[f'A{row}']
                    cell.value = 'פירוט לפי תקופות'
                    cell.fill = section_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    row += 1
                    
                    # Period breakdown headers
                    headers = ['תקופה', 'תאריך התחלה', 'תאריך סיום', 'הכנסות', 'הוצאות', 'רווח']
                    for idx, header in enumerate(headers, 1):
                        col = get_column_letter(idx)
                        cell = ws[f'{col}{row}']
                        cell.value = header
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    row += 1
                    
                    # Period breakdown data
                    for period in all_year_periods:
                        period_label = f"תקופה {period.year_index}" if period.year_index > 1 else "תקופה ראשית"
                        ws[f'A{row}'] = period_label
                        ws[f'B{row}'] = period.start_date.isoformat()
                        ws[f'C{row}'] = period.end_date.isoformat()
                        ws[f'D{row}'] = float(period.total_income or 0)
                        ws[f'D{row}'].number_format = '#,##0.00'
                        ws[f'E{row}'] = float(period.total_expense or 0)
                        ws[f'E{row}'].number_format = '#,##0.00'
                        ws[f'F{row}'] = float(period.total_profit or 0)
                        ws[f'F{row}'].number_format = '#,##0.00'
                        for col_idx in range(1, 7):
                            col = get_column_letter(col_idx)
                            ws[f'{col}{row}'].border = border
                        row += 1
                    row += 1
            
            # Budgets
            if summary.get('budgets'):
                ws.merge_cells(f'A{row}:F{row}')
                cell = ws[f'A{row}']
                cell.value = 'תקציבים'
                cell.fill = section_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                # Budget headers
                headers = ['קטגוריה', 'סכום (₪)', 'סוג תקופה', 'תאריך התחלה', 'תאריך סיום', 'פעיל']
                for idx, header in enumerate(headers, 1):
                    col = get_column_letter(idx)
                    cell = ws[f'{col}{row}']
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                # Budget data
                for budget in summary['budgets']:
                    ws[f'A{row}'] = budget.get('category', '')
                    ws[f'B{row}'] = budget.get('amount', 0)
                    ws[f'B{row}'].number_format = '#,##0.00'
                    
                    p_type = budget.get('period_type', '')
                    if p_type == 'Annual':
                        p_type = 'שנתי'
                    elif p_type == 'Monthly':
                        p_type = 'חודשי'
                    ws[f'C{row}'] = p_type
                    
                    ws[f'D{row}'] = budget.get('start_date', '') or ''
                    ws[f'E{row}'] = budget.get('end_date', '') or ''
                    ws[f'F{row}'] = 'כן' if budget.get('is_active', False) else 'לא'
                    for col_idx in range(1, 7):
                        col = get_column_letter(col_idx)
                        ws[f'{col}{row}'].border = border
                    row += 1
                    row += 1
            
            # Fund Transactions
            if summary.get('fund_transactions'):
                ws.merge_cells(f'A{row}:G{row}')
                cell = ws[f'A{row}']
                cell.value = 'עסקאות קופה'
                cell.fill = section_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                # Fund Transaction headers
                headers = ['תאריך', 'סוג', 'סכום (₪)', 'תיאור', 'קטגוריה', 'אמצעי תשלום', 'הערות']
                for idx, header in enumerate(headers, 1):
                    col = get_column_letter(idx)
                    cell = ws[f'{col}{row}']
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                # Fund Transaction data
                for tx in summary['fund_transactions']:
                    tx_type = 'הכנסה' if tx.get('type') == 'Income' else 'הוצאה'
                    amount = tx.get('amount', 0)
                    
                    ws[f'A{row}'] = tx.get('tx_date', '')
                    ws[f'B{row}'] = tx_type
                    ws[f'C{row}'] = amount
                    ws[f'C{row}'].number_format = '#,##0.00'
                    ws[f'D{row}'] = tx.get('description', '') or ''
                    ws[f'E{row}'] = tx.get('category', '') or ''
                    ws[f'F{row}'] = tx.get('payment_method', '') or ''
                    ws[f'G{row}'] = tx.get('notes', '') or ''
                    
                    # Color code by type
                    if tx.get('type') == 'Income':
                        ws[f'B{row}'].fill = income_fill
                        ws[f'C{row}'].fill = income_fill
                    else:
                        ws[f'B{row}'].fill = expense_fill
                        ws[f'C{row}'].fill = expense_fill
                    
                    for col_idx in range(1, 8):
                        col = get_column_letter(col_idx)
                        ws[f'{col}{row}'].border = border
                    row += 1
                row += 1
            
            # Transactions
            if summary.get('transactions'):
                ws.merge_cells(f'A{row}:G{row}')
                cell = ws[f'A{row}']
                cell.value = 'עסקאות'
                cell.fill = section_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                # Transaction headers
                headers = ['תאריך', 'סוג', 'סכום (₪)', 'תיאור', 'קטגוריה', 'אמצעי תשלום', 'הערות']
                for idx, header in enumerate(headers, 1):
                    col = get_column_letter(idx)
                    cell = ws[f'{col}{row}']
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                # Transaction data
                for tx in summary['transactions']:
                    tx_type = 'הכנסה' if tx.get('type') == 'Income' else 'הוצאה'
                    amount = tx.get('amount', 0)
                    
                    ws[f'A{row}'] = tx.get('tx_date', '')
                    ws[f'B{row}'] = tx_type
                    ws[f'C{row}'] = amount
                    ws[f'C{row}'].number_format = '#,##0.00'
                    ws[f'D{row}'] = tx.get('description', '') or ''
                    ws[f'E{row}'] = tx.get('category', '') or ''
                    ws[f'F{row}'] = tx.get('payment_method', '') or ''
                    ws[f'G{row}'] = tx.get('notes', '') or ''
                    
                    # Color code by type
                    if tx.get('type') == 'Income':
                        ws[f'B{row}'].fill = income_fill
                        ws[f'C{row}'].fill = income_fill
                    else:
                        ws[f'B{row}'].fill = expense_fill
                        ws[f'C{row}'].fill = expense_fill
                    
                    for col_idx in range(1, 8):
                        col = get_column_letter(col_idx)
                        ws[f'{col}{row}'].border = border
                    row += 1
            
            # Auto-adjust column widths
            for col in ws.columns:
                try:
                    if not col:
                        continue
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except Exception:
                            logger.debug("Could not read cell value for column width calculation")
                    adjusted_width = min(max_length + 2, 50)
                    if adjusted_width > 0:
                        ws.column_dimensions[col_letter].width = adjusted_width
                except Exception as e:
                    # Skip if there's an error adjusting column width
                    logger.warning("Could not adjust column width: %s", e)
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Create filename - use ASCII-safe version to avoid encoding issues in headers
            import re
            from datetime import datetime as dt
            # Remove non-ASCII characters from filename for header compatibility
            safe_project_name = re.sub(r'[^\x00-\x7F]', '_', project.name).replace('"', '').replace('/', '_').replace('\\', '_').strip()
            
            # Add date range to filename
            start_date = summary.get('start_date', '')
            end_date = summary.get('end_date', '')
            
            if start_date and end_date:
                # Parse dates and format for filename
                try:
                    if isinstance(start_date, str):
                        start_dt = dt.fromisoformat(start_date.split('T')[0])
                    else:
                        start_dt = start_date
                    if isinstance(end_date, str):
                        end_dt = dt.fromisoformat(end_date.split('T')[0])
                    else:
                        end_dt = end_date
                    date_range = f"{start_dt.strftime('%Y-%m-%d')}_{end_dt.strftime('%Y-%m-%d')}"
                except (ValueError, AttributeError):
                    logger.debug("Could not parse date range for filename, using raw values")
                    date_range = f"{start_date}_{end_date}"
            elif start_date:
                try:
                    if isinstance(start_date, str):
                        start_dt = dt.fromisoformat(start_date.split('T')[0])
                    else:
                        start_dt = start_date
                    date_range = f"מ-{start_dt.strftime('%Y-%m-%d')}"
                except (ValueError, AttributeError):
                    logger.debug("Could not parse start_date for filename, using raw value")
                    date_range = f"מ-{start_date}"
            elif end_date:
                try:
                    if isinstance(end_date, str):
                        end_dt = dt.fromisoformat(end_date.split('T')[0])
                    else:
                        end_dt = end_date
                    date_range = f"עד-{end_dt.strftime('%Y-%m-%d')}"
                except (ValueError, AttributeError):
                    logger.debug("Could not parse end_date for filename, using raw value")
                    date_range = f"עד-{end_date}"
            else:
                date_range = "כל-התקופות"
            
            filename = f"{safe_project_name}_{date_range}.xlsx"
            
            return Response(
                content=output.getvalue(),
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': f'attachment; filename="{filename}"'
                }
            )
        except ImportError as import_err:
            # Fallback to CSV if openpyxl is not available
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header with Hebrew support
            writer.writerow(['סיכום תקופת חוזה'])
            writer.writerow([])
            writer.writerow(['שם פרויקט', project.name])
            writer.writerow(['שנת חוזה', summary['year_label']])
            writer.writerow(['תאריך התחלה', summary['start_date']])
            writer.writerow(['תאריך סיום', summary['end_date']])
            writer.writerow([])
            writer.writerow(['סיכום כלכלי'])
            writer.writerow(['סה"כ הכנסות', summary['total_income']])
            writer.writerow(['סה"כ הוצאות', summary['total_expense']])
            writer.writerow(['סה"כ רווח', summary['total_profit']])
            writer.writerow([])
            
            # Check if this is a previous year period and add year summary if so
            from datetime import date
            current_year = date.today().year
            contract_year = summary.get('contract_year', current_year)
            
            # If this is a previous year (not current year), add summary of all periods in that year
            if contract_year < current_year:
                from backend.repositories.contract_period_repository import ContractPeriodRepository
                period_repo = ContractPeriodRepository(db)
                all_year_periods = await period_repo.get_by_project_and_year(project_id, contract_year)
                
                if len(all_year_periods) > 0:  # Show if there are any periods
                    # Calculate year totals
                    year_total_income = sum(float(p.total_income or 0) for p in all_year_periods)
                    year_total_expense = sum(float(p.total_expense or 0) for p in all_year_periods)
                    year_total_profit = year_total_income - year_total_expense
                    
                    writer.writerow([])
                    writer.writerow(['סיכום כולל של כל התקופות בשנה'])
                    writer.writerow(['שנה', contract_year])
                    writer.writerow(['מספר תקופות', len(all_year_periods)])
                    writer.writerow([])
                    writer.writerow(['סיכום כלכלי כולל של השנה'])
                    writer.writerow(['סה"כ הכנסות', year_total_income])
                    writer.writerow(['סה"כ הוצאות', year_total_expense])
                    writer.writerow(['סה"כ רווח', year_total_profit])
                    writer.writerow([])
                    
                    # Add breakdown by period
                    writer.writerow(['פירוט לפי תקופות'])
                    writer.writerow(['תקופה', 'תאריך התחלה', 'תאריך סיום', 'הכנסות', 'הוצאות', 'רווח'])
                    for period in all_year_periods:
                        period_label = f"תקופה {period.year_index}" if period.year_index > 1 else "תקופה ראשית"
                        writer.writerow([
                            period_label,
                            period.start_date.isoformat(),
                            period.end_date.isoformat(),
                            float(period.total_income or 0),
                            float(period.total_expense or 0),
                            float(period.total_profit or 0)
                        ])
                    writer.writerow([])
            
            # Write budgets
            if summary.get('budgets'):
                writer.writerow(['תקציבים'])
                writer.writerow(['קטגוריה', 'סכום', 'סוג תקופה', 'תאריך התחלה', 'תאריך סיום', 'פעיל'])
                for budget in summary['budgets']:
                    p_type = budget.get('period_type', '')
                    if p_type == 'Annual':
                        p_type = 'שנתי'
                    elif p_type == 'Monthly':
                        p_type = 'חודשי'
                    writer.writerow([
                        budget.get('category', ''),
                        budget.get('amount', 0),
                        p_type,
                        budget.get('start_date', ''),
                        budget.get('end_date', ''),
                        'כן' if budget.get('is_active', False) else 'לא'
                    ])
                writer.writerow([])
            
            # Write transactions
            if summary.get('transactions'):
                writer.writerow(['עסקאות'])
                writer.writerow([
                    'תאריך',
                    'סוג',
                    'סכום',
                    'תיאור',
                    'קטגוריה',
                    'אמצעי תשלום',
                    'הערות'
                ])
                
                for tx in summary['transactions']:
                    writer.writerow([
                        tx.get('tx_date', ''),
                        'הכנסה' if tx.get('type') == 'Income' else 'הוצאה',
                        tx.get('amount', 0),
                        tx.get('description', '') or '',
                        tx.get('category', '') or '',
                        tx.get('payment_method', '') or '',
                        tx.get('notes', '') or ''
                    ])
            
            # Prepare response with UTF-8 BOM for Excel compatibility
            csv_content = output.getvalue()
            output.close()
            
            # Add BOM for proper Hebrew display in Excel
            csv_bytes = '\ufeff'.encode('utf-8') + csv_content.encode('utf-8-sig')
            
            # Create filename - use ASCII-safe version to avoid encoding issues in headers
            import re
            # Remove non-ASCII characters from filename for header compatibility
            safe_project_name = re.sub(r'[^\x00-\x7F]', '_', project.name).replace('"', '').replace('/', '_').replace('\\', '_').strip()
            safe_year_label = re.sub(r'[^\x00-\x7F]', '_', str(summary["year_label"])).replace('"', '').replace('/', '_').replace('\\', '_').strip()
            filename = f"contract_period_{safe_year_label}_{safe_project_name}.csv"
            
            return Response(
                content=csv_bytes,
                media_type='text/csv; charset=utf-8',
                headers={
                    'Content-Disposition': f'attachment; filename="{filename}"'
                }
            )
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Error exporting contract period CSV/Excel")
        raise HTTPException(status_code=500, detail="שגיאה בייצוא הקובץ")


@router.post("/{project_id}/check-contract-renewal")
async def check_and_renew_contract(
    project_id: int,
    db: DBSessionDep,
    user = Depends(get_current_user)
):
    """Check if contract has ended and renew it automatically if needed"""
    service = ContractPeriodService(db)
    renewed_project = await service.check_and_renew_contract(project_id)
    
    if renewed_project:
        # Reload contract periods after renewal to ensure they're up to date
        periods_by_year = await service.get_previous_contracts_by_year(project_id)
        result = []
        for year in sorted(periods_by_year.keys(), reverse=False):
            result.append({
                'year': year,
                'periods': periods_by_year[year]
            })
        
        return {
            'renewed': True,
            'message': 'חוזה חודש בהצלחה',
            'new_start_date': renewed_project.start_date.isoformat() if renewed_project.start_date else None,
            'new_end_date': renewed_project.end_date.isoformat() if renewed_project.end_date else None,
            'contract_periods': {
                'project_id': project_id,
                'periods_by_year': result
            }
        }
    else:
        return {
            'renewed': False,
            'message': 'החוזה עדיין לא הסתיים או אין תאריך סיום מוגדר'
        }


@router.get("/{project_id}/contract-periods/year/{year}/export-csv")
async def export_contract_year_csv(
    project_id: int,
    year: int,
    db: DBSessionDep,
    user = Depends(get_current_user)
):
    """Export all contract periods for a specific year to Excel"""
    try:
        service = ContractPeriodService(db)
        from backend.repositories.contract_period_repository import ContractPeriodRepository
        period_repo = ContractPeriodRepository(db)
        all_year_periods = await period_repo.get_by_project_and_year(project_id, year)
        
        # Get project info
        project = await ProjectRepository(db).get_by_id(project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        # If no periods exist for this year, create a virtual one for the whole year
        if not all_year_periods:
            from backend.models.contract_period import ContractPeriod
            from datetime import date as date_type
            
            # Determine start and end dates for the year, clipped by project dates
            year_start = max(project.start_date, date_type(year, 1, 1)) if project.start_date else date_type(year, 1, 1)
            project_end = project.end_date if project.end_date else date_type.today()
            year_end = min(project_end, date_type(year, 12, 31))
            
            if year_start > year_end:
                raise HTTPException(status_code=404, detail=f"Project does not cover year {year}")
                
            # Create a virtual period object (not saved to DB)
            virtual_period = ContractPeriod(
                id=None,
                project_id=project_id,
                start_date=year_start,
                end_date=year_end,
                contract_year=year,
                year_index=1
            )
            
            # Use the virtual period for summary
            summary_data = await service.get_contract_period_summary(
                project_id=project_id,
                start_date=year_start,
                end_date=year_end
            )
            
            # Reconstruct all_year_periods structure for the rest of the function
            # We'll use a simplified flow for virtual year
            summary = {
                'year_label': f"שנת {year}",
                'start_date': year_start.isoformat(),
                'end_date': year_end.isoformat(),
                'total_income': summary_data.get('total_income', 0),
                'total_expense': summary_data.get('total_expense', 0),
                'total_profit': summary_data.get('total_profit', 0),
                'transactions': summary_data.get('transactions', []),
                'fund_transactions': summary_data.get('fund_transactions', []),
                'budgets': summary_data.get('budgets', []),
                'all_periods': [virtual_period],
                'contract_year': year
            }
        else:
            # Sort periods by start date
            all_year_periods.sort(key=lambda p: p.start_date)
            
            # Aggregate data for the whole year
            year_total_income = 0
            year_total_expense = 0
            all_transactions = []
            all_fund_transactions = []
            all_budgets = []
            
            for period in all_year_periods:
                summary_data = await service.get_contract_period_summary(period.id)
                year_total_income += summary_data.get('total_income', 0)
                year_total_expense += summary_data.get('total_expense', 0)
                
                # Add period label to transactions to distinguish them
                period_label = summary_data.get('year_label') or f"תקופה {period.year_index}"
                for tx in summary_data.get('transactions', []):
                    tx['period_label'] = period_label
                    all_transactions.append(tx)
                for tx in summary_data.get('fund_transactions', []):
                    tx['period_label'] = period_label
                    all_fund_transactions.append(tx)
                    
                # Use budgets from the latest period in that year as they are likely the most up to date
                all_budgets = summary_data.get('budgets', [])

            summary = {
                'year_label': f"שנת {year}",
                'start_date': all_year_periods[0].start_date.isoformat(),
                'end_date': all_year_periods[-1].end_date.isoformat() if all_year_periods[-1].end_date else date.today().isoformat(),
                'total_income': year_total_income,
                'total_expense': year_total_expense,
                'total_profit': year_total_income - year_total_expense,
                'transactions': all_transactions,
                'fund_transactions': all_fund_transactions,
                'budgets': all_budgets,
                'all_periods': all_year_periods,
                'contract_year': year
            }

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"סיכום שנת {year}"
            ws.sheet_view.rightToLeft = True
            ws.sheet_view.rightToLeft = True
            
            # Define colors and styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            income_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            expense_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            profit_positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            profit_negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            header_font = Font(bold=True, color="FFFFFF", size=12)
            title_font = Font(bold=True, color="FFFFFF", size=14)
            normal_font = Font(size=11)
            bold_font = Font(bold=True, size=11)
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            row = 1
            
            # Title
            ws.merge_cells(f'A{row}:B{row}')
            cell = ws[f'A{row}']
            cell.value = f'סיכום שנת חוזה {year}'
            cell.fill = title_fill
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 2
            
            # Project info
            ws[f'A{row}'] = 'שם פרויקט'
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = project.name
            row += 1
            
            ws[f'A{row}'] = 'שנה'
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = year
            row += 1
            
            ws[f'A{row}'] = 'תאריך התחלה'
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = summary['start_date']
            row += 1
            
            ws[f'A{row}'] = 'תאריך סיום'
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = summary['end_date']
            row += 2
            
            # Financial Summary
            ws.merge_cells(f'A{row}:B{row}')
            cell = ws[f'A{row}']
            cell.value = 'סיכום כלכלי שנתי'
            cell.fill = section_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            ws[f'A{row}'] = 'סוג'
            ws[f'B{row}'] = 'סכום (₪)'
            for col in ['A', 'B']:
                cell = ws[f'{col}{row}']
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            ws[f'A{row}'] = 'סה"כ הכנסות'
            ws[f'B{row}'] = summary['total_income']
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'].fill = income_fill
            ws[f'B{row}'].number_format = '#,##0.00'
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = border
            row += 1
            
            ws[f'A{row}'] = 'סה"כ הוצאות'
            ws[f'B{row}'] = summary['total_expense']
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'].fill = expense_fill
            ws[f'B{row}'].number_format = '#,##0.00'
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = border
            row += 1
            
            ws[f'A{row}'] = 'סה"כ רווח'
            ws[f'B{row}'] = summary['total_profit']
            ws[f'A{row}'].font = bold_font
            profit_fill = profit_positive_fill if summary['total_profit'] >= 0 else profit_negative_fill
            ws[f'B{row}'].fill = profit_fill
            ws[f'B{row}'].number_format = '#,##0.00'
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = border
            row += 2

            # Periods Summary Table
            ws.merge_cells(f'A{row}:F{row}')
            cell = ws[f'A{row}']
            cell.value = 'פירוט תקופות בשנה'
            cell.fill = section_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            headers = ['תקופה', 'תאריך התחלה', 'תאריך סיום', 'הכנסות', 'הוצאות', 'רווח']
            for idx, header in enumerate(headers, 1):
                col = get_column_letter(idx)
                cell = ws[f'{col}{row}']
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            for p in all_year_periods:
                p_label = f"תקופה {p.year_index}"
                ws[f'A{row}'] = p_label
                ws[f'B{row}'] = p.start_date.isoformat()
                ws[f'C{row}'] = p.end_date.isoformat() if p.end_date else ''
                ws[f'D{row}'] = float(p.total_income or 0)
                ws[f'D{row}'].number_format = '#,##0.00'
                ws[f'E{row}'] = float(p.total_expense or 0)
                ws[f'E{row}'].number_format = '#,##0.00'
                ws[f'F{row}'] = float(p.total_profit or 0)
                ws[f'F{row}'].number_format = '#,##0.00'
                for col_idx in range(1, 7):
                    col = get_column_letter(col_idx)
                    ws[f'{col}{row}'].border = border
                row += 1
            row += 2

            # Budgets
            if summary.get('budgets'):
                ws.merge_cells(f'A{row}:F{row}')
                cell = ws[f'A{row}']
                cell.value = 'תקציבים'
                cell.fill = section_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                headers = ['קטגוריה', 'סכום (₪)', 'סוג תקופה', 'תאריך התחלה', 'תאריך סיום', 'פעיל']
                for idx, header in enumerate(headers, 1):
                    col = get_column_letter(idx)
                    cell = ws[f'{col}{row}']
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                for budget in summary['budgets']:
                    ws[f'A{row}'] = budget.get('category', '')
                    ws[f'B{row}'] = budget.get('amount', 0)
                    ws[f'B{row}'].number_format = '#,##0.00'
                    p_type = budget.get('period_type', '')
                    if p_type == 'Annual':
                        p_type = 'שנתי'
                    elif p_type == 'Monthly':
                        p_type = 'חודשי'
                    ws[f'C{row}'] = p_type
                    ws[f'D{row}'] = budget.get('start_date', '') or ''
                    ws[f'E{row}'] = budget.get('end_date', '') or ''
                    ws[f'F{row}'] = 'כן' if budget.get('is_active', False) else 'לא'
                    for col_idx in range(1, 7):
                        col = get_column_letter(col_idx)
                        ws[f'{col}{row}'].border = border
                    row += 1
                row += 2

            # Fund Transactions
            if summary.get('fund_transactions'):
                ws.merge_cells(f'A{row}:H{row}')
                cell = ws[f'A{row}']
                cell.value = 'עסקאות קופה'
                cell.fill = section_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                headers = ['תאריך', 'תקופה', 'סוג', 'סכום (₪)', 'תיאור', 'קטגוריה', 'אמצעי תשלום', 'הערות']
                for idx, header in enumerate(headers, 1):
                    col = get_column_letter(idx)
                    cell = ws[f'{col}{row}']
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                for tx in summary['fund_transactions']:
                    tx_type = 'הכנסה' if tx.get('type') == 'Income' else 'הוצאה'
                    ws[f'A{row}'] = tx.get('tx_date', '')
                    ws[f'B{row}'] = tx.get('period_label', '')
                    ws[f'C{row}'] = tx_type
                    ws[f'D{row}'] = tx.get('amount', 0)
                    ws[f'D{row}'].number_format = '#,##0.00'
                    ws[f'E{row}'] = tx.get('description', '') or ''
                    ws[f'F{row}'] = tx.get('category', '') or ''
                    ws[f'G{row}'] = tx.get('payment_method', '') or ''
                    ws[f'H{row}'] = tx.get('notes', '') or ''
                    
                    if tx.get('type') == 'Income':
                        ws[f'C{row}'].fill = income_fill
                        ws[f'D{row}'].fill = income_fill
                    else:
                        ws[f'C{row}'].fill = expense_fill
                        ws[f'D{row}'].fill = expense_fill
                    
                    for col_idx in range(1, 9):
                        col = get_column_letter(col_idx)
                        ws[f'{col}{row}'].border = border
                    row += 1
                row += 2

            # All Transactions
            if summary.get('transactions'):
                ws.merge_cells(f'A{row}:H{row}')
                cell = ws[f'A{row}']
                cell.value = 'עסקאות'
                cell.fill = section_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                headers = ['תאריך', 'תקופה', 'סוג', 'סכום (₪)', 'תיאור', 'קטגוריה', 'אמצעי תשלום', 'הערות']
                for idx, header in enumerate(headers, 1):
                    col = get_column_letter(idx)
                    cell = ws[f'{col}{row}']
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                for tx in summary['transactions']:
                    tx_type = 'הכנסה' if tx.get('type') == 'Income' else 'הוצאה'
                    ws[f'A{row}'] = tx.get('tx_date', '')
                    ws[f'B{row}'] = tx.get('period_label', '')
                    ws[f'C{row}'] = tx_type
                    ws[f'D{row}'] = tx.get('amount', 0)
                    ws[f'D{row}'].number_format = '#,##0.00'
                    ws[f'E{row}'] = tx.get('description', '') or ''
                    ws[f'F{row}'] = tx.get('category', '') or ''
                    ws[f'G{row}'] = tx.get('payment_method', '') or ''
                    ws[f'H{row}'] = tx.get('notes', '') or ''
                    
                    if tx.get('type') == 'Income':
                        ws[f'C{row}'].fill = income_fill
                        ws[f'D{row}'].fill = income_fill
                    else:
                        ws[f'C{row}'].fill = expense_fill
                        ws[f'D{row}'].fill = expense_fill
                    
                    for col_idx in range(1, 9):
                        col = get_column_letter(col_idx)
                        ws[f'{col}{row}'].border = border
                    row += 1

            # Auto-adjust column widths
            for col in ws.columns:
                try:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = min(max_length + 2, 50)
                    if adjusted_width > 0:
                        ws.column_dimensions[col_letter].width = adjusted_width
                except Exception:
                    logger.debug("Could not adjust column width in year export Excel")

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            import re
            safe_project_name = re.sub(r'[^\x00-\x7F]', '_', project.name).replace('"', '').replace('/', '_').replace('\\', '_').strip()
            filename = f"{safe_project_name}_year_{year}.xlsx"
            
            return Response(
                content=output.getvalue(),
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': f'attachment; filename="{filename}"'
                }
            )
        except ImportError:
            # Simple CSV fallback
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([f'סיכום שנת {year}', project.name])
            writer.writerow(['סה"כ הכנסות', summary['total_income']])
            writer.writerow(['סה"כ הוצאות', summary['total_expense']])
            writer.writerow(['סה"כ רווח', summary['total_profit']])
            writer.writerow([])
            writer.writerow(['עסקאות'])
            writer.writerow(['תאריך', 'תקופה', 'סוג', 'סכום', 'תיאור'])
            for tx in summary['transactions']:
                tx_type = 'הכנסה' if tx.get('type') == 'Income' else 'הוצאה'
                writer.writerow([tx.get('tx_date'), tx.get('period_label'), tx_type, tx.get('amount'), tx.get('description')])
            
            return Response(
                content=output.getvalue(),
                media_type='text/csv',
                headers={
                    'Content-Disposition': f'attachment; filename="{year}_summary.csv"'
                }
            )
    except Exception as e:
        logger.exception("Error exporting year CSV")
        raise HTTPException(status_code=500, detail="שגיאה בייצוא הקובץ")


@router.post("/{project_id}/close-year")
async def close_contract_year(
    project_id: int,
    db: DBSessionDep,
    end_date: str = Form(..., description="End date in YYYY-MM-DD format"),
    user = Depends(require_permission("delete", "project", resource_id_param="project_id", project_id_param=None))
):
    """
    Manually close a contract year and archive it.
    This creates a read-only archive entry and starts a new contract period.
    """
    try:
        # Parse date string to date object
        from datetime import datetime as dt
        end_date_obj = dt.strptime(end_date, "%Y-%m-%d").date()
        
        service = ContractPeriodService(db)
        contract_period = await service.close_year_manually(
            project_id=project_id,
            end_date=end_date_obj,
            archived_by_user_id=user.id
        )
        
        # Ensure recurring transactions are generated for the new period
        from backend.services.recurring_transaction_service import RecurringTransactionService
        recurring_service = RecurringTransactionService(db)
        await recurring_service.ensure_project_transactions_generated(project_id)
        
        # Reload contract periods after closing
        periods_by_year = await service.get_previous_contracts_by_year(project_id)
        
        return {
            'success': True,
            'message': 'שנה נסגרה ונשמרה בארכיון בהצלחה',
            'contract_period_id': contract_period.id,
            'start_date': contract_period.start_date.isoformat(),
            'end_date': contract_period.end_date.isoformat(),
            'contract_year': contract_period.contract_year,
            'periods_by_year': periods_by_year
        }
    except ValueError as e:
        logger.exception("Validation error in close_contract_year for project %s", project_id)
        raise HTTPException(status_code=400, detail="שגיאה בסגירת שנת חוזה")